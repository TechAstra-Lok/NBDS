import csv
import io
import re
from urllib.parse import urljoin, urlparse
from flask import (
    Blueprint, render_template, request, redirect,
    url_for, flash, jsonify, session, Response
)
from flask_login import login_user, logout_user, login_required, current_user
from app import db
from app.models import (
    User, Donor, BloodRequest, News, Notice,
    Advertisement, Contact, SiteVisitor, SuccessStory, StaffMember, Partner, BloodBank, BloodInventory, BloodInventoryMovement, BloodReservation, BloodTransfer, LowStockAlert, Notification, AuditLog, Volunteer, NotificationDeliveryLog
)
from app.utils import generate_qr_code
from app.services.auth_service import AuthService
from app.forms import (
    AdminLoginForm, DonorAdminCreateForm, DonorEditForm,
    NewsForm, NoticeForm, AdvertisementForm, AdminUserForm,
    StaffMemberForm, PartnerForm
)
from app.utils import save_image, save_file, delete_file, paginate_query, sanitize_html
from sqlalchemy import desc, func, or_  # यहाँ or_ इम्पोर्ट फिक्स गरिएको छ
from datetime import datetime, timedelta
from functools import wraps
from werkzeug.security import generate_password_hash

admin_bp = Blueprint('admin', __name__)


# Enforce short admin session (5 minutes) based on last activity
@admin_bp.before_request
def enforce_admin_session_timeout():
    # Only apply to logged-in users accessing admin blueprint
    from datetime import datetime
    if not current_user.is_authenticated:
        return

    last = session.get('admin_last_active')
    if last:
        try:
            last_dt = datetime.fromisoformat(last)
        except Exception:
            last_dt = None
        if last_dt:
            if datetime.utcnow() - last_dt > timedelta(minutes=5):
                # expire session
                logout_user()
                session.pop('admin_last_active', None)
                flash('Your admin session has expired due to inactivity. Please log in again.', 'warning')
                return redirect(url_for('admin.login'))

    # update last active timestamp for admins
    session['admin_last_active'] = datetime.utcnow().isoformat()


# ─── Auth Decorators ──────────────────────────
def superadmin_required(f):
    @wraps(f)
    @login_required
    def decorated(*args, **kwargs):
        if not current_user.is_superadmin:
            flash('Super Admin access required.', 'danger')
            return redirect(url_for('admin.dashboard'))
        return f(*args, **kwargs)
    return decorated


from app.rbac import permission_required


def _resolve_bank_tenant(bank):
    """
    If the blood bank has been provisioned with a tenant database,
    resolve its tenant context so that queries against tenant-scoped
    models (BloodInventory, etc.) hit the correct DB file.
    Safe to call on un-provisioned banks (no-op).
    """
    if bank.tenant_id and bank.db_name and bank.tenant_status == 'Active':
        try:
            from app.services.tenant_service import TenantResolutionService
            TenantResolutionService.resolve_tenant(bank.tenant_id)
        except Exception:
            pass  # fall back to main DB


def build_blood_bank_dashboard_summary(bank_id):
    bank = BloodBank.query.get_or_404(bank_id)
    _resolve_bank_tenant(bank)
    inventory_items = BloodInventory.query.filter_by(blood_bank_id=bank.id).all()
    low_stock_items = [item for item in inventory_items if item.available_units < item.minimum_stock]
    pending_transfers = BloodTransfer.query.filter_by(destination_bank_id=bank.id, status='pending').count()
    alerts = LowStockAlert.query.filter_by(blood_bank_id=bank.id).count()
    critical_items = sum(1 for item in inventory_items if item.available_units <= max(1, item.minimum_stock // 2))
    return {
        'bank_id': bank.id,
        'bank_name': bank.display_name,
        'inventory_count': len(inventory_items),
        'low_stock_count': len(low_stock_items),
        'pending_transfers': pending_transfers,
        'alert_count': alerts,
        'critical_items': critical_items,
    }


def create_inventory_notifications(inventory):
    notifications = []
    if inventory.available_units < inventory.minimum_stock:
        notification = Notification(
            title='Low stock alert',
            message=f"{inventory.blood_group} {inventory.component} is below minimum stock.",
            category='low_stock',
            channel='in_app',
        )
        db.session.add(notification)
        notifications.append(notification)

    if inventory.expiry_date:
        from datetime import datetime
        try:
            expiry = datetime.strptime(inventory.expiry_date, '%Y-%m-%d').date()
        except ValueError:
            expiry = None
        if expiry and (expiry - datetime.utcnow().date()).days <= 14:
            notification = Notification(
                title='Expiry soon',
                message=f"{inventory.blood_group} {inventory.component} expires soon.",
                category='expiry',
                channel='in_app',
            )
            db.session.add(notification)
            notifications.append(notification)
    return notifications


def log_audit_event(action, entity_id, details, actor='system'):
    audit_entry = AuditLog(action=action, entity_id=entity_id, details=details, actor=actor)
    db.session.add(audit_entry)
    return audit_entry


def build_blood_inventory_report(bank_id):
    bank = BloodBank.query.get_or_404(bank_id)
    _resolve_bank_tenant(bank)
    inventory_items = BloodInventory.query.filter_by(blood_bank_id=bank.id).all()
    low_stock_count = sum(1 for item in inventory_items if item.available_units < item.minimum_stock)
    expiring_soon_count = 0
    for item in inventory_items:
        if item.expiry_date:
            from datetime import datetime
            try:
                expiry = datetime.strptime(item.expiry_date, '%Y-%m-%d').date()
            except ValueError:
                expiry = None
            if expiry and (expiry - datetime.utcnow().date()).days <= 14:
                expiring_soon_count += 1
    return {
        'bank_id': bank.id,
        'bank_name': bank.display_name,
        'inventory_count': len(inventory_items),
        'low_stock_count': low_stock_count,
        'expiring_soon_count': expiring_soon_count,
        'pending_transfers': BloodTransfer.query.filter_by(destination_bank_id=bank.id, status='pending').count(),
    }


# ════════════════════════════════════════════
#   LOGIN / LOGOUT
# ════════════════════════════════════════════
@admin_bp.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('admin.dashboard'))
    
    form = AdminLoginForm()
    
    if form.validate_on_submit():
        user = User.query.filter_by(username=form.username.data).first()
        
        if user and user.is_active and user.check_password(form.password.data):
            login_user(user, remember=form.remember.data)
            # make admin session permanent for session lifetime tracking
            session.permanent = True
            session['admin_last_active'] = datetime.utcnow().isoformat()
            user.last_login = datetime.utcnow()
            db.session.commit()
            
            next_page = request.args.get('next')
            if next_page and not is_safe_url(next_page):
                next_page = None
            flash(f'✅ Welcome back, {user.full_name or user.username}!', 'success')
            return redirect(next_page or url_for('admin.dashboard'))
        
        flash('❌ Invalid credentials. Please try again.', 'danger')
    
    return render_template('admin/login.html', form=form)


def is_safe_url(target):
    if not target:
        return False
    ref_url = urlparse(request.host_url)
    test_url = urlparse(urljoin(request.host_url, target))
    return test_url.scheme in ('http', 'https') and ref_url.netloc == test_url.netloc


@admin_bp.route('/logout')
@login_required
def logout():
    logout_user()
    flash('You have been logged out.', 'info')
    return redirect(url_for('public.index'))


# ════════════════════════════════════════════
#   DASHBOARD
# ════════════════════════════════════════════
@admin_bp.route('/')
@admin_bp.route('/dashboard')
@login_required
def dashboard():
    # Core stats
    total_donors    = Donor.query.count()
    avail_donors    = Donor.query.filter_by(availability_status='available').count()
    recently_donated_donors = Donor.query.filter_by(availability_status='recently_donated').count()
    unavailable_donors = Donor.query.filter_by(availability_status='unavailable').count()
    
    total_requests  = BloodRequest.query.count()
    active_requests = BloodRequest.query.filter_by(status='active').count()
    fulfilled       = BloodRequest.query.filter_by(status='fulfilled').count()
    
    total_news      = News.query.filter_by(is_published=True).count()
    total_notices   = Notice.query.filter_by(is_active=True).count()
    unread_contacts = Contact.query.filter_by(is_read=False).count()
    total_stories   = SuccessStory.query.count()
    pending_success_stories = SuccessStory.query.filter_by(status='pending').count()
    
    total_partners  = Partner.query.count()
    total_staff     = StaffMember.query.count()
    total_advertisements = Advertisement.query.count()
    
    volunteer_doctors = Volunteer.query.filter_by(designation='Doctor').count()
    volunteer_nurses  = Volunteer.query.filter_by(designation='Nurse').count()
    volunteer_has     = Volunteer.query.filter_by(designation='HA').count()
    
    total_blood_banks = BloodBank.query.count()
    pending_events    = News.query.filter(News.category.in_(['event', 'program'])).filter(News.scheduled_date > datetime.utcnow()).count()
    
    notifications_sent = NotificationDeliveryLog.query.filter_by(status='sent').count()
    notifications_failed = NotificationDeliveryLog.query.filter_by(status='failed').count()

    # Visitor stats
    today       = datetime.utcnow().date()
    today_visitors = SiteVisitor.query.filter_by(visit_date=today).count()
    week_ago    = today - timedelta(days=7)
    week_visitors = SiteVisitor.query.filter(SiteVisitor.visit_date >= week_ago).count()
    total_visitors = SiteVisitor.query.count()
    
    # Blood group breakdown
    bg_breakdown = db.session.query(
        Donor.blood_group,
        func.count(Donor.id).label('count')
    ).group_by(Donor.blood_group).order_by(desc('count')).all()
    
    # Recent activity
    recent_donors   = Donor.query.order_by(desc(Donor.created_at)).limit(5).all()
    recent_requests = BloodRequest.query.order_by(desc(BloodRequest.created_at)).limit(5).all()
    
    # Monthly registration trend (last 6 months)
    monthly_data = []
    for i in range(5, -1, -1):
        month_start = (datetime.utcnow().replace(day=1) - timedelta(days=i*30)).replace(day=1)
        month_end   = (month_start + timedelta(days=32)).replace(day=1)
        count = Donor.query.filter(
            Donor.created_at >= month_start,
            Donor.created_at < month_end
        ).count()
        monthly_data.append({
            'month': month_start.strftime('%b %Y'),
            'count': count
        })

    # District coverage and emergency support count
    districts_covered = db.session.query(func.count(func.distinct(Donor.curr_district))).scalar() or 0
    active_emergencies = BloodRequest.query.filter_by(status='active', is_emergency=True).count()
    
    return render_template('admin/dashboard.html',
        total_donors=total_donors,
        avail_donors=avail_donors,
        recently_donated_donors=recently_donated_donors,
        unavailable_donors=unavailable_donors,
        total_requests=total_requests,
        active_requests=active_requests,
        fulfilled=fulfilled,
        total_news=total_news,
        total_notices=total_notices,
        unread_contacts=unread_contacts,
        total_stories=total_stories,
        pending_success_stories=pending_success_stories,
        total_partners=total_partners,
        total_staff=total_staff,
        total_advertisements=total_advertisements,
        volunteer_doctors=volunteer_doctors,
        volunteer_nurses=volunteer_nurses,
        volunteer_has=volunteer_has,
        total_blood_banks=total_blood_banks,
        pending_events=pending_events,
        notifications_sent=notifications_sent,
        notifications_failed=notifications_failed,
        today_visitors=today_visitors,
        week_visitors=week_visitors,
        total_visitors=total_visitors,
        bg_breakdown=bg_breakdown,
        recent_donors=recent_donors,
        recent_requests=recent_requests,
        monthly_data=monthly_data,
        districts_covered=districts_covered,
        active_emergencies=active_emergencies,
    )


# ════════════════════════════════════════════
#   BLOOD BANK MANAGEMENT
# ════════════════════════════════════════════



@admin_bp.route('/blood-banks/<int:bank_id>/inventory', methods=['GET', 'POST'])
@permission_required('manage_blood_banks')
def blood_bank_inventory(bank_id):
    bank = BloodBank.query.get_or_404(bank_id)
    _resolve_bank_tenant(bank)
    if request.method == 'POST':
        inventory = BloodInventory(
            blood_bank_id=bank.id,
            blood_group=request.form.get('blood_group', '').strip(),
            component=request.form.get('component', 'Whole Blood').strip() or 'Whole Blood',
            units_available=int(request.form.get('units_available', 0) or 0),
            units_reserved=int(request.form.get('units_reserved', 0) or 0),
            minimum_stock=int(request.form.get('minimum_stock', 4) or 4),
            maximum_stock=int(request.form.get('maximum_stock', 20) or 20),
            expiry_date=request.form.get('expiry_date', '').strip() or None,
        )
        db.session.add(inventory)
        db.session.flush()
        inventory.qr_code = generate_qr_code('inventory', inventory.id)
        db.session.commit()
        if inventory.units_available < inventory.minimum_stock:
            db.session.add(LowStockAlert(
                blood_bank_id=bank.id,
                blood_group=inventory.blood_group,
                component=inventory.component,
                severity='warning',
                message=f"{inventory.blood_group} {inventory.component} stock is below minimum level.",
            ))
        create_inventory_notifications(inventory)
        log_audit_event('inventory_created', inventory.id, f"Inventory created for {inventory.blood_group}/{inventory.component}", actor='admin')
        db.session.commit()
        if request.form.get('movement_type') and request.form.get('movement_units'):
            movement = BloodInventoryMovement(
                inventory_id=inventory.id,
                movement_type=request.form.get('movement_type', 'received').strip(),
                units=int(request.form.get('movement_units', 0) or 0),
                note=request.form.get('movement_note', '').strip() or None,
            )
            db.session.add(movement)
            db.session.commit()
        flash('Inventory entry added.', 'success')
        return redirect(url_for('admin.blood_bank_inventory', bank_id=bank.id))

    inventory_items = BloodInventory.query.filter_by(blood_bank_id=bank.id).order_by(BloodInventory.blood_group).all()
    inventory_map = {item.id: item.movements for item in inventory_items}
    reservations = BloodReservation.query.filter_by(blood_bank_id=bank.id).order_by(BloodReservation.requested_at.desc()).all()
    transfers = BloodTransfer.query.filter((BloodTransfer.source_bank_id == bank.id) | (BloodTransfer.destination_bank_id == bank.id)).order_by(BloodTransfer.created_at.desc()).all()
    alerts = LowStockAlert.query.filter_by(blood_bank_id=bank.id).order_by(LowStockAlert.created_at.desc()).all()
    blood_banks = BloodBank.query.filter(BloodBank.id != bank.id).order_by(BloodBank.name).all()

    summary = build_blood_bank_dashboard_summary(bank.id)
    report = build_blood_inventory_report(bank.id)
    return render_template('admin/blood_bank_inventory.html', bank=bank, inventory_items=inventory_items, reservations=reservations, transfers=transfers, alerts=alerts, blood_banks=blood_banks, inventory_map=inventory_map, dashboard_summary=summary, report_summary=report)


@admin_bp.route('/blood-banks/<int:bank_id>/transfers', methods=['POST'])
@permission_required('manage_blood_banks')
def blood_bank_transfers(bank_id):
    bank = BloodBank.query.get_or_404(bank_id)
    _resolve_bank_tenant(bank)
    transfer = BloodTransfer(
        source_bank_id=bank.id,
        destination_bank_id=int(request.form.get('destination_bank_id') or 0),
        blood_group=request.form.get('blood_group', '').strip(),
        component=request.form.get('component', 'Whole Blood').strip() or 'Whole Blood',
        units=int(request.form.get('units', 0) or 0),
        status='pending',
        remarks=request.form.get('remarks', '').strip() or None,
    )
    db.session.add(transfer)
    db.session.commit()
    flash('Transfer request recorded.', 'success')
    return redirect(url_for('admin.blood_bank_inventory', bank_id=bank.id))


# ════════════════════════════════════════════
#   DONOR MANAGEMENT
# ════════════════════════════════════════════
@admin_bp.route('/donors')
@permission_required('manage_donors')
def donors():
    page        = request.args.get('page', 1, type=int)
    search      = request.args.get('q', '')
    blood_group = request.args.get('bg', '')
    status      = request.args.get('status', '')
    donor_type  = request.args.get('type', '')
    
    query = Donor.query
    
    if search:
        query = query.filter(or_(
            Donor.full_name.ilike(f'%{search}%'),
            Donor.donor_id.ilike(f'%{search}%'),
            Donor.phone1.ilike(f'%{search}%'),
            Donor.curr_district.ilike(f'%{search}%'),
        ))
    if blood_group:
        query = query.filter_by(blood_group=blood_group)
    if status:
        query = query.filter_by(availability_status=status)
    if donor_type:
        query = query.filter_by(donor_type=donor_type)
    
    pagination = paginate_query(
        query.order_by(desc(Donor.created_at)), page, 20
    )
    
    total = Donor.query.count()
    blood_groups = ['A+', 'A-', 'B+', 'B-', 'AB+', 'AB-', 'O+', 'O-']
    
    return render_template('admin/donors.html',
        pagination=pagination,
        total=total,
        blood_groups=blood_groups,
        search=search,
        selected_bg=blood_group,
        selected_status=status,
        selected_type=donor_type,
    )


@admin_bp.route('/donors/export')
@admin_bp.route('/donors/export-csv')
@permission_required('manage_donors')
def export_donors():
    """Export all or filtered donors as CSV, Excel (.xlsx), or PDF (.pdf)."""
    fmt = request.args.get('format', request.args.get('fmt', 'csv')).lower()
    search      = request.args.get('q', '')
    blood_group = request.args.get('bg', '')
    status      = request.args.get('status', '')
    donor_type  = request.args.get('type', '')
    
    query = Donor.query
    if search:
        query = query.filter(or_(
            Donor.full_name.ilike(f'%{search}%'),
            Donor.donor_id.ilike(f'%{search}%'),
            Donor.phone1.ilike(f'%{search}%'),
            Donor.curr_district.ilike(f'%{search}%'),
        ))
    if blood_group:
        query = query.filter_by(blood_group=blood_group)
    if status:
        query = query.filter_by(availability_status=status)
    if donor_type:
        query = query.filter_by(donor_type=donor_type)
        
    donors_list = query.order_by(desc(Donor.created_at)).all()
    
    HEADERS = [
        'Donor ID', 'Full Name', 'Email', 'Primary Phone', 'Secondary Phone',
        'Blood Group', 'Age', 'Gender', 'Weight (kg)', 'Donor Type',
        'Availability Status', 'Last Donation Date', 'Current Province',
        'Current District', 'Current Local Level', 'Current Ward', 'Current Tole',
        'Permanent Province', 'Permanent District', 'Permanent Local Level',
        'Registered Date'
    ]

    rows = []
    for d in donors_list:
        rows.append([
            d.donor_id or '',
            d.full_name or '',
            d.email or '',
            d.phone1 or '',
            d.phone2 or '',
            d.blood_group or '',
            str(d.age) if d.age else '',
            d.gender or '',
            str(d.weight) if d.weight else '',
            d.donor_type or 'regular',
            d.availability_status or 'available',
            d.last_donation_date.strftime('%Y-%m-%d') if d.last_donation_date else '',
            d.curr_province or '',
            d.curr_district or '',
            d.curr_local_level or '',
            d.curr_ward or '',
            d.curr_tole or '',
            d.perm_province or '',
            d.perm_district or '',
            d.perm_local_level or '',
            d.created_at.strftime('%Y-%m-%d %H:%M:%S') if d.created_at else ''
        ])

    if fmt in ('xlsx', 'excel'):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            flash('openpyxl is required for Excel export.', 'danger')
            return redirect(url_for('admin.donors'))

        wb = Workbook()
        ws = wb.active
        ws.title = 'Blood Donors'

        header_fill = PatternFill('solid', fgColor='DC2626')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        thin = Side(style='thin', color='D1D5DB')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        ws.append(HEADERS)
        for col_idx in range(1, len(HEADERS) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
        ws.row_dimensions[1].height = 30

        for r_data in rows:
            ws.append(r_data)
            r = ws.max_row
            for col_idx in range(1, len(HEADERS) + 1):
                ws.cell(row=r, column=col_idx).border = border

        for col_idx, header in enumerate(HEADERS, 1):
            max_len = max((len(str(r[col_idx-1])) for r in rows), default=0)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max(len(header), max_len) + 3, 35)

        ws.freeze_panes = 'A2'
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return Response(
            buf.read(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': 'attachment; filename=nepal_blood_donors.xlsx'}
        )

    if fmt == 'pdf':
        try:
            from reportlab.lib.pagesizes import landscape, A4
            from reportlab.lib import colors
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.pdfgen import canvas
        except ImportError:
            flash('reportlab is required for PDF export.', 'danger')
            return redirect(url_for('admin.donors'))

        class NumberedCanvas(canvas.Canvas):
            def __init__(self, *args, **kwargs):
                super().__init__(*args, **kwargs)
                self._saved_page_states = []
            def showPage(self):
                self._saved_page_states.append(dict(self.__dict__))
                self._startPage()
            def save(self):
                num_pages = len(self._saved_page_states)
                for state in self._saved_page_states:
                    self.__dict__.update(state)
                    self.draw_page_number(num_pages)
                    super().showPage()
                super().save()
            def draw_page_number(self, page_count):
                self.saveState()
                self.setFont('Helvetica', 8)
                self.setFillColor(colors.HexColor('#6B7280'))
                self.drawRightString(self._pagesize[0] - 30, 20, f'Page {self._pageNumber} of {page_count}')
                self.drawString(30, 20, 'Raktadata — Nepali Blood Donors Society | Registered Donors Export')
                self.setStrokeColor(colors.HexColor('#E5E7EB'))
                self.setLineWidth(0.5)
                self.line(30, 32, self._pagesize[0] - 30, 32)
                self.restoreState()

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=landscape(A4), rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=40)
        styles = getSampleStyleSheet()

        title_style = ParagraphStyle('DocTitle', parent=styles['Heading1'], fontName='Helvetica-Bold', fontSize=15, textColor=colors.HexColor('#DC2626'), spaceAfter=4)
        sub_style = ParagraphStyle('DocSub', parent=styles['Normal'], fontName='Helvetica', fontSize=9, textColor=colors.HexColor('#4B5563'), spaceAfter=10)
        cell_style = ParagraphStyle('CellText', parent=styles['Normal'], fontName='Helvetica', fontSize=8, leading=10, textColor=colors.HexColor('#1F2937'))
        header_cell_style = ParagraphStyle('HeaderCellText', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=9, leading=11, textColor=colors.whitesmoke)

        filters_used = []
        if search: filters_used.append(f'Search: "{search}"')
        if blood_group: filters_used.append(f'Group: {blood_group}')
        if status: filters_used.append(f'Status: {status}')
        if donor_type: filters_used.append(f'Type: {donor_type}')
        filter_str = (' | '.join(filters_used)) if filters_used else 'All Donors'
        now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        elements = [
            Paragraph('NEPALI BLOOD DONORS SOCIETY — REGISTERED DONORS DIRECTORY', title_style),
            Paragraph(f'Exported on: {now_str} | Filters: {filter_str} | Total Records: {len(donors_list)}', sub_style),
            Spacer(1, 4)
        ]

        pdf_headers = ['Donor ID', 'Full Name', 'Blood Group', 'Phone', 'Location', 'Donor Type', 'Status', 'Last Donated']
        table_data = [[Paragraph(h, header_cell_style) for h in pdf_headers]]

        for d in donors_list:
            loc = f"{d.curr_district or 'N/A'}"
            if d.curr_local_level:
                loc += f", {d.curr_local_level}"
            ld_date = d.last_donation_date.strftime('%Y-%m-%d') if d.last_donation_date else 'N/A'
            table_data.append([
                Paragraph(d.donor_id or 'N/A', cell_style),
                Paragraph(d.full_name or 'N/A', cell_style),
                Paragraph(d.blood_group or 'N/A', cell_style),
                Paragraph(d.phone1 or 'N/A', cell_style),
                Paragraph(loc, cell_style),
                Paragraph((d.donor_type or 'regular').capitalize(), cell_style),
                Paragraph((d.availability_status or 'available').capitalize(), cell_style),
                Paragraph(ld_date, cell_style),
            ])

        col_widths = [65, 140, 60, 95, 150, 80, 80, 90]
        t = Table(table_data, colWidths=col_widths, repeatRows=1)
        t_style = [
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#DC2626')),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('BOTTOMPADDING', (0,0), (-1,0), 6),
            ('TOPPADDING', (0,0), (-1,0), 6),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E5E7EB')),
        ]
        for r_i in range(1, len(table_data)):
            bg = colors.HexColor('#F9FAFB') if r_i % 2 == 0 else colors.white
            t_style.append(('BACKGROUND', (0, r_i), (-1, r_i), bg))
            t_style.append(('TOPPADDING', (0, r_i), (-1, r_i), 4))
            t_style.append(('BOTTOMPADDING', (0, r_i), (-1, r_i), 4))

        t.setStyle(TableStyle(t_style))
        elements.append(t)
        doc.build(elements, canvasmaker=NumberedCanvas)

        buf.seek(0)
        return Response(
            buf.read(),
            mimetype='application/pdf',
            headers={'Content-Disposition': 'attachment; filename=nepal_blood_donors.pdf'}
        )

    # Default CSV
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(HEADERS)
    writer.writerows(rows)
    output.seek(0)
    return Response(
        output.getvalue(),
        mimetype='text/csv',
        headers={'Content-Disposition': 'attachment; filename=nepal_blood_donors_export.csv'}
    )


@admin_bp.route('/donors/sample-csv')
@permission_required('manage_donors')
def sample_donors_csv():
    """Generate and return a sample CSV template for bulk donor upload."""
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        'Full Name', 'Current Age', 'Current Weight', 'Permanent Address',
        'Current Address', 'Phone number 1', 'Phone number 2', 'Blood Group',
        'Previous Blood Donation Date (if any)', 'Previous Blood Donation Location',
        'Type of Donor', 'Social Medial Profile Link(e.g Facebook, Instagram)'
    ])
    writer.writerow([
        'Ram Bahadur Shrestha', '28', '65', 'Kavre',
        'Kathmandu', '9841234567', '9801234567', 'O+',
        '2025-10-15', 'Bir Hospital',
        'regular', 'https://facebook.com/ram'
    ])
    writer.writerow([
        'Sita Kumari Thapa', '24', '55', 'Kaski',
        'Pokhara', '9841987654', '', 'A+',
        '', '',
        'emergency', 'https://instagram.com/sita'
    ])
    output.seek(0)
    return Response(
        output.getvalue(),
        mimetype='text/csv',
        headers={'Content-Disposition': 'attachment; filename=sample_blood_donors_import.csv'}
    )


@admin_bp.route('/donors/sample-excel')
@permission_required('manage_donors')
def sample_donors_excel():
    """Generate and return a sample Excel (.xlsx) template for bulk donor upload."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        flash('openpyxl is required.', 'danger')
        return redirect(url_for('admin.donors'))

    HEADERS = [
        'Full Name', 'Current Age', 'Current Weight', 'Permanent Address',
        'Current Address', 'Phone number 1', 'Phone number 2', 'Blood Group',
        'Previous Blood Donation Date (if any)', 'Previous Blood Donation Location',
        'Type of Donor', 'Social Medial Profile Link(e.g Facebook, Instagram)'
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = 'Sample Donor Import'
    ws.append(HEADERS)

    header_fill = PatternFill('solid', fgColor='DC2626')
    header_font = Font(bold=True, color='FFFFFF')
    for col_idx in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    ws.append([
        'Ram Bahadur Shrestha', '28', '65', 'Kavre',
        'Kathmandu', '9841234567', '9801234567', 'O+',
        '2025-10-15', 'Bir Hospital',
        'regular', 'https://facebook.com/ram'
    ])
    ws.append([
        'Sita Kumari Thapa', '24', '55', 'Kaski',
        'Pokhara', '9841987654', '', 'A+',
        '', '',
        'emergency', 'https://instagram.com/sita'
    ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        buf.read(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment; filename=sample_blood_donors_import.xlsx'}
    )


@admin_bp.route('/donors/import-csv', methods=['POST'])
@admin_bp.route('/donors/bulk-import', methods=['POST'])
@permission_required('manage_donors')
def import_donors_csv():
    """Bulk import blood donors from a CSV or Excel (.xlsx) file with duplicate handling (skip or override)."""
    file_field = 'csv_file' if 'csv_file' in request.files else ('excel_file' if 'excel_file' in request.files else 'file')
    if file_field not in request.files:
        flash('No file uploaded.', 'danger')
        return redirect(url_for('admin.donors'))
        
    file = request.files[file_field]
    if not file or file.filename == '':
        flash('No file selected for upload.', 'danger')
        return redirect(url_for('admin.donors'))
        
    filename = file.filename.lower()
    if not (filename.endswith('.csv') or filename.endswith('.xlsx') or filename.endswith('.xls')):
        flash('Invalid file format. Please upload a .csv or .xlsx file.', 'danger')
        return redirect(url_for('admin.donors'))
        
    duplicate_action = request.form.get('duplicate_action', 'skip').strip().lower()
    
    try:
        raw_bytes = file.stream.read()
        rows_data = []

        if filename.endswith('.xlsx') or filename.endswith('.xls'):
            try:
                from openpyxl import load_workbook
            except ImportError:
                flash('openpyxl is required for Excel files.', 'danger')
                return redirect(url_for('admin.donors'))

            wb = load_workbook(io.BytesIO(raw_bytes), read_only=True, data_only=True)
            ws = wb.active
            raw_grid = list(ws.values)
            if not raw_grid:
                flash('Uploaded file is empty.', 'danger')
                return redirect(url_for('admin.donors'))

            headers = [str(h).strip() if h is not None else '' for h in raw_grid[0]]
            for r_vals in raw_grid[1:]:
                row_dict = {}
                for h_col, val in zip(headers, r_vals):
                    if h_col:
                        row_dict[h_col] = str(val).strip() if val is not None else ''
                rows_data.append(row_dict)
        else:
            # CSV parsing
            stream = io.StringIO(raw_bytes.decode('utf-8-sig', errors='replace'))
            sample = stream.read(8192)
            stream.seek(0)
            try:
                dialect = csv.Sniffer().sniff(sample, delimiters=',\t;|')
                csv_reader = csv.DictReader(stream, dialect=dialect)
            except csv.Error:
                csv_reader = csv.DictReader(stream)
            rows_data = list(csv_reader)  # fallback to comma
        
        imported_count = 0
        updated_count = 0
        skipped_count = 0
        skipped_reasons = []
        
        # DEBUG: Log detected headers and delimiter for troubleshooting
        import logging
        _log = logging.getLogger('csv_import')
        detected_delim = repr(dialect.delimiter) if 'dialect' in dir() else 'comma(default)'
        raw_headers = csv_reader.fieldnames or []
        _log.warning(f"CSV IMPORT DEBUG: delimiter={detected_delim}, headers={raw_headers}")
        _log.warning(f"CSV IMPORT DEBUG: raw_bytes first 500 chars: {repr(raw_bytes[:500])}")
        
        default_pin_hash = generate_password_hash('1234')
        
        FIELD_ALIASES = {
            # Full Name
            'full name': 'full_name', 'full_name': 'full_name', 'name': 'full_name',
            'donor name': 'full_name', 'donor_name': 'full_name', 'fullname': 'full_name',
            # Email
            'email': 'email', 'email address': 'email', 'email_address': 'email',
            'e-mail': 'email', 'emailaddress': 'email',
            # Primary Phone
            'primary phone': 'phone1', 'phone1': 'phone1', 'phone number 1': 'phone1',
            'phone': 'phone1', 'mobile': 'phone1', 'contact': 'phone1',
            'phone number': 'phone1', 'mobile number': 'phone1', 'contact number': 'phone1',
            'primary phone number': 'phone1', 'primaryphone': 'phone1',
            # Secondary Phone
            'secondary phone': 'phone2', 'phone2': 'phone2', 'phone number 2': 'phone2',
            'secondary phone number': 'phone2', 'secondaryphone': 'phone2',
            'alternate phone': 'phone2', 'alt phone': 'phone2',
            # Blood Group
            'blood group': 'blood_group', 'blood_group': 'blood_group',
            'bloodgroup': 'blood_group', 'bg': 'blood_group', 'blood type': 'blood_group',
            'blood_type': 'blood_group', 'bloodtype': 'blood_group',
            # Age
            'age': 'age', 'current age': 'age', 'current_age': 'age', 'donor age': 'age',
            # Gender
            'gender': 'gender', 'sex': 'gender',
            # Weight
            'weight': 'weight', 'weight (kg)': 'weight', 'weight(kg)': 'weight',
            'current weight': 'weight', 'current_weight': 'weight', 'weight_kg': 'weight',
            # Donor Type
            'donor type': 'donor_type', 'donor_type': 'donor_type', 'type': 'donor_type',
            'type of donor': 'donor_type', 'donortype': 'donor_type',
            # Availability Status
            'availability status': 'availability_status', 'availability_status': 'availability_status',
            'status': 'availability_status', 'availability': 'availability_status',
            # Last Donation Date
            'last donation date': 'last_donation_date', 'last_donation_date': 'last_donation_date',
            'last donation': 'last_donation_date', 'last_donation': 'last_donation_date',
            'previous blood donation date (if any)': 'last_donation_date',
            'previous blood donation date': 'last_donation_date',
            'previous blood donated date(last time)': 'last_donation_date',
            'previous blood donated date': 'last_donation_date',
            'donation date': 'last_donation_date', 'last donated': 'last_donation_date',
            # Donation Times
            'donation_times': 'donation_times', 'donation times': 'donation_times',
            'previous blood donated times(e.g 1, 2 , ... if not write \'0\')': 'donation_times',
            'previous blood donated times': 'donation_times',
            # Current Address fields
            'current province': 'curr_province', 'curr_province': 'curr_province',
            'province': 'curr_province', 'current_province': 'curr_province',
            'current district': 'curr_district', 'curr_district': 'curr_district',
            'district': 'curr_district', 'current_district': 'curr_district',
            'current local level': 'curr_local_level', 'curr_local_level': 'curr_local_level',
            'local level': 'curr_local_level', 'current_local_level': 'curr_local_level',
            'municipality': 'curr_local_level', 'city': 'curr_local_level',
            'current ward': 'curr_ward', 'curr_ward': 'curr_ward', 'ward': 'curr_ward',
            'current_ward': 'curr_ward',
            'current tole': 'curr_tole', 'curr_tole': 'curr_tole', 'tole': 'curr_tole',
            'current_tole': 'curr_tole',
            'current address': 'current_address', 'current_address': 'current_address',
            # Permanent Address fields
            'permanent province': 'perm_province', 'perm_province': 'perm_province',
            'permanent_province': 'perm_province',
            'permanent district': 'perm_district', 'perm_district': 'perm_district',
            'permanent_district': 'perm_district',
            'permanent local level': 'perm_local_level', 'perm_local_level': 'perm_local_level',
            'permanent_local_level': 'perm_local_level',
            'permanent ward': 'perm_ward', 'perm_ward': 'perm_ward',
            'permanent_ward': 'perm_ward',
            'permanent tole': 'perm_tole', 'perm_tole': 'perm_tole',
            'permanent_tole': 'perm_tole',
            'permanent address': 'permanent_address', 'permanent_address': 'permanent_address',
            # Social Link
            'social medial profile link(e.g facebook, instagram)': 'social_link',
            'social link': 'social_link', 'social_link': 'social_link',
            'social media': 'social_link', 'facebook': 'social_link', 'instagram': 'social_link',
            'social profile': 'social_link', 'social media link': 'social_link',
            # Registered Date
            'registered date': 'registered_date', 'registered_date': 'registered_date',
            'registration date': 'registered_date', 'created at': 'registered_date',
            # Donor ID
            'donor id': '_donor_id', 'donor_id': '_donor_id', 'id': '_donor_id',
        }
        
        def _normalize_row(raw_row):
            normalized = {}
            for raw_key, value in raw_row.items():
                if not raw_key:
                    continue
                clean_key = raw_key.strip().lower()
                canonical = FIELD_ALIASES.get(clean_key)
                if canonical and canonical not in normalized:
                    normalized[canonical] = value.strip() if value else ''
                elif not canonical and clean_key not in normalized:
                    normalized[clean_key] = value.strip() if value else ''
            return normalized

        VALID_BLOOD_GROUPS = ['A+', 'A-', 'B+', 'B-', 'AB+', 'AB-', 'O+', 'O-']
        BG_NAME_MAP = {
            'AB POSITIVE': 'AB+', 'AB NEGATIVE': 'AB-',
            'AB+VE': 'AB+', 'AB-VE': 'AB-',
            'A POSITIVE': 'A+', 'A NEGATIVE': 'A-',
            'A+VE': 'A+', 'A-VE': 'A-',
            'B POSITIVE': 'B+', 'B NEGATIVE': 'B-',
            'B+VE': 'B+', 'B-VE': 'B-',
            'O POSITIVE': 'O+', 'O NEGATIVE': 'O-',
            'O+VE': 'O+', 'O-VE': 'O-',
            'A POS': 'A+', 'A NEG': 'A-',
            'B POS': 'B+', 'B NEG': 'B-',
            'AB POS': 'AB+', 'AB NEG': 'AB-',
            'O POS': 'O+', 'O NEG': 'O-',
        }

        def _parse_blood_group(raw_val):
            if not raw_val:
                return 'O+'
            raw_str = raw_val.strip().upper()
            candidates = []
            if '(' in raw_str:
                parts = raw_str.split('(')
                before_p = parts[0].strip()
                after_p = parts[1].replace(')', '').strip() if len(parts) > 1 else ''
                if before_p: candidates.append(before_p)
                if after_p: candidates.append(after_p)
            else:
                candidates.append(raw_str)
            
            for cand in candidates:
                if cand in VALID_BLOOD_GROUPS:
                    return cand
                if cand in BG_NAME_MAP:
                    return BG_NAME_MAP[cand]
                cand_no_space = cand.replace(' ', '')
                if cand_no_space.endswith('VE') and len(cand_no_space) > 2:
                    cand_no_space = cand_no_space[:-2]
                if cand_no_space in VALID_BLOOD_GROUPS:
                    return cand_no_space
            return 'O+'

        for idx, row in enumerate(rows_data, start=2):
            try:
                with db.session.begin_nested():
                    row_data = _normalize_row(row)
                    
                    full_name = row_data.get('full_name', '').strip()
                    if not full_name:
                        # Fallback search in raw row values for non-empty text string
                        for k, v in row.items():
                            if v and isinstance(v, str) and len(v.strip()) > 1 and not v.strip().replace('.','',1).isdigit():
                                full_name = v.strip()
                                break
                    
                    if not full_name:
                        skipped_count += 1
                        skipped_reasons.append(f"Row {idx}: missing Full Name")
                        continue
                    
                    phone1 = row_data.get('phone1', '').strip()
                    email = row_data.get('email', '').strip()
                    blood_group_raw = row_data.get('blood_group', '').strip()
                    
                    # DEBUG: Log first 3 rows for troubleshooting
                    if idx <= 4:
                        _log.warning(f"CSV ROW {idx} DEBUG: blood_group_raw={repr(blood_group_raw)}, keys={list(row_data.keys())}")
                        _log.warning(f"CSV ROW {idx} RAW: {dict(row)}")
                    
                    # Ensure phone1 is present or generate unique placeholder
                    if not phone1:
                        generated_p = f"9000{idx:06d}"
                        phone1 = generated_p
                    
                    # Ensure email is present or generate unique placeholder
                    if not email:
                        email = f"donor_{phone1}@nepaliblooddonors.org"
                    
                    bg_clean = _parse_blood_group(blood_group_raw)
                    
                    # DEBUG: Log parsed blood group for first 3 rows
                    if idx <= 4:
                        _log.warning(f"CSV ROW {idx} PARSED: bg_clean={repr(bg_clean)}")
                    
                    # Check existing donor by phone or email
                    existing = Donor.query.filter(
                        (Donor.phone1 == phone1) | (Donor.email == email)
                    ).first()

                    # Age
                    try:
                        age_val = row_data.get('age', '').strip()
                        age = int(float(age_val)) if age_val else 25
                    except (ValueError, TypeError):
                        age = 25
                    
                    # Weight
                    try:
                        weight_val = row_data.get('weight', '').strip()
                        weight = float(weight_val) if weight_val else 60.0
                    except (ValueError, TypeError):
                        weight = 60.0
                    
                    # Gender
                    gender = row_data.get('gender', '').strip().lower()
                    if gender not in ['male', 'female', 'other', 'prefer_not_to_say']:
                        gender = 'male'
                    
                    # Donor type
                    donor_type = row_data.get('donor_type', '').strip().lower()
                    if 'occ' in donor_type or 'reg' in donor_type or 'norm' in donor_type:
                        donor_type = 'regular'
                    elif 'vol' in donor_type:
                        donor_type = 'volunteer'
                    elif 'emerg' in donor_type or 'urg' in donor_type:
                        donor_type = 'emergency'
                    elif 'plat' in donor_type or 'sdp' in donor_type:
                        donor_type = 'platelet'
                    elif 'rare' in donor_type:
                        donor_type = 'rare'
                    elif donor_type not in ['regular', 'emergency', 'platelet', 'rare', 'volunteer', 'other']:
                        donor_type = donor_type[:30] if donor_type else 'regular'
                    
                    # Availability status
                    avail_status = row_data.get('availability_status', '').strip().lower()
                    if avail_status not in ['available', 'recently_donated', 'unavailable']:
                        avail_status = 'available'
                    
                    # Last donation date
                    last_donation_date = None
                    ld_str = row_data.get('last_donation_date', '').strip()
                    if ld_str:
                        # Split to remove time if only date is needed, but strptime with time is better
                        ld_clean = ld_str.split(' ')[0] if ' ' in ld_str and not ':' in ld_str else ld_str
                        for fmt in ('%m/%d/%Y %H:%M:%S', '%m/%d/%Y %H:%M', '%m/%d/%Y', '%Y-%m-%d', '%d/%m/%Y', '%Y/%m/%d',
                                    '%d-%m-%Y', '%m-%d-%Y', '%B %d, %Y', '%d %B %Y', '%Y-%m-%d %H:%M:%S', '%Y-%m-%dT%H:%M:%S'):
                            try:
                                last_donation_date = datetime.strptime(ld_clean, fmt).date()
                                break
                            except ValueError:
                                pass
                    
                    # Address fields
                    curr_addr_fallback = row_data.get('current_address', '').strip()
                    
                    def _infer_province(addr_text):
                        if not addr_text: return 'Bagmati'
                        text = addr_text.lower()
                        prov_map = {
                            'Koshi': ['jhapa', 'morang', 'sunsari', 'dhankuta', 'birtamode', 'arjundhara', 'damak', 'dharan', 'biratnagar'],
                            'Madhesh': ['saptari', 'siraha', 'dhanusha', 'janakpur', 'birgunj'],
                            'Bagmati': ['kathmandu', 'lalitpur', 'bhaktapur', 'chitwan'],
                            'Gandaki': ['kaski', 'pokhara', 'gorkha'],
                            'Lumbini': ['rupandehi', 'butwal', 'dang'],
                            'Karnali': ['surkhet', 'birendranagar'],
                            'Sudurpashchim': ['kailali', 'dhangadhi']
                        }
                        for p, kws in prov_map.items():
                            if any(k in text for k in kws): return p
                        return 'Bagmati'

                    curr_province = row_data.get('curr_province', '').strip()
                    if not curr_province:
                        curr_province = _infer_province(curr_addr_fallback)
                        
                    curr_district = row_data.get('curr_district', '').strip() or curr_addr_fallback or 'Kathmandu'
                    curr_local_level = row_data.get('curr_local_level', '').strip() or curr_addr_fallback or 'Kathmandu'
                    curr_ward = row_data.get('curr_ward', '').strip()
                    curr_tole = row_data.get('curr_tole', '').strip()
                    
                    perm_addr_fallback = row_data.get('permanent_address', '').strip()
                    perm_province = row_data.get('perm_province', '').strip()
                    if not perm_province:
                        perm_province = _infer_province(perm_addr_fallback) if perm_addr_fallback else curr_province
                        
                    perm_district = row_data.get('perm_district', '').strip() or perm_addr_fallback or curr_district
                    perm_local_level = row_data.get('perm_local_level', '').strip() or perm_addr_fallback or curr_local_level
                    perm_ward = row_data.get('perm_ward', '').strip()
                    perm_tole = row_data.get('perm_tole', '').strip()
                    
                    social_link = row_data.get('social_link', '').strip()
                    phone2 = row_data.get('phone2', '').strip()
                    
                    donation_times = 0
                    dt_str = row_data.get('donation_times', '').strip()
                    if dt_str:

                        match = re.search(r'\d+', dt_str)
                        if match:
                            donation_times = int(match.group())
                    
                    created_at_dt = None
                    reg_str = row_data.get('registered_date', '').strip()
                    if reg_str:
                        for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', '%Y/%m/%d',
                                    '%d-%m-%Y', '%m-%d-%Y', '%B %d, %Y', '%d %B %Y',
                                    '%Y-%m-%d %H:%M:%S', '%Y-%m-%dT%H:%M:%S'):
                            try:
                                created_at_dt = datetime.strptime(reg_str, fmt)
                                break
                            except ValueError:
                                pass

                    if existing:
                        if duplicate_action in ['override', 'update']:
                            # Update existing donor record
                            existing.full_name = full_name
                            if phone2: existing.phone2 = phone2
                            if age: existing.age = age
                            if weight: existing.weight = weight
                            if bg_clean: existing.blood_group = bg_clean
                            if gender: existing.gender = gender
                            if donor_type: existing.donor_type = donor_type
                            if avail_status: existing.availability_status = avail_status
                            if curr_province: existing.curr_province = curr_province
                            if curr_district: existing.curr_district = curr_district
                            if curr_local_level: existing.curr_local_level = curr_local_level
                            if curr_ward: existing.curr_ward = curr_ward
                            if curr_tole: existing.curr_tole = curr_tole
                            if perm_province: existing.perm_province = perm_province
                            if perm_district: existing.perm_district = perm_district
                            if perm_local_level: existing.perm_local_level = perm_local_level
                            if perm_ward: existing.perm_ward = perm_ward
                            if perm_tole: existing.perm_tole = perm_tole
                            if last_donation_date: existing.last_donation_date = last_donation_date
                            if donation_times: existing.donation_times = donation_times
                            if social_link: existing.social_link = social_link
                            
                            existing.recalculate_and_save()
                            updated_count += 1
                        else:
                            skipped_count += 1
                            skipped_reasons.append(f"Row {idx}: duplicate ({full_name})")
                    else:
                        # Create new donor record
                        donor = Donor(
                            full_name=full_name,
                            email=email,
                            phone1=phone1,
                            phone2=phone2,
                            pin_hash=default_pin_hash,
                            age=age,
                            weight=weight,
                            blood_group=bg_clean,
                            gender=gender,
                            donor_type=donor_type,
                            availability_status=avail_status,
                            curr_province=curr_province,
                            curr_district=curr_district,
                            curr_local_level=curr_local_level,
                            curr_ward=curr_ward,
                            curr_tole=curr_tole,
                            perm_province=perm_province,
                            perm_district=perm_district,
                            perm_local_level=perm_local_level,
                            perm_ward=perm_ward,
                            perm_tole=perm_tole,
                            last_donation_date=last_donation_date,
                            donation_times=donation_times,
                            social_link=social_link,
                            is_active=True,
                            is_public=True
                        )
                        if created_at_dt:
                            donor.created_at = created_at_dt
                        donor.recalculate_and_save()
                        db.session.add(donor)
                        imported_count += 1

            except Exception as row_err:
                skipped_count += 1
                skipped_reasons.append(f"Row {idx}: error ({str(row_err)})")
            
        db.session.commit()
        
        audit_log = AuditLog(
            action='BULK_IMPORT_DONORS',
            details=f'Bulk import completed. Added: {imported_count}, Updated: {updated_count}, Skipped: {skipped_count}.',
            actor=current_user.username if hasattr(current_user, 'username') else 'admin'
        )
        db.session.add(audit_log)
        db.session.commit()
        
        msg_parts = []
        if imported_count > 0:
            msg_parts.append(f"✅ Imported {imported_count} new donors")
        if updated_count > 0:
            msg_parts.append(f"🔄 Updated {updated_count} existing donors")
        if skipped_count > 0:
            msg_parts.append(f"⚠️ {skipped_count} rows skipped")
            
        msg = ". ".join(msg_parts) if msg_parts else "No donors processed."
        if skipped_reasons:
            detail = "; ".join(skipped_reasons[:5])
            if len(skipped_reasons) > 5:
                detail += f" ... and {len(skipped_reasons) - 5} more"
            msg += f" (Reasons: {detail})"
            
        flash(msg, 'success' if (imported_count > 0 or updated_count > 0) else 'warning')
        
    except Exception as e:
        db.session.rollback()
        flash(f"❌ Error processing CSV file: {str(e)}", 'danger')
        
    return redirect(url_for('admin.donors'))


@admin_bp.route('/donors/add', methods=['GET', 'POST'])
@permission_required('manage_donors')
def add_donor():
    form = DonorAdminCreateForm()
    
    if request.method == 'POST':
        if form.validate_on_submit():
            try:
                email_val = form.email.data.strip() if form.email.data and form.email.data.strip() else None
                donor = Donor(
                    full_name           = form.full_name.data.strip(),
                    email               = email_val,
                    pin_hash            = generate_password_hash('1234'),
                    age                 = form.age.data,
                    weight              = form.weight.data,
                    perm_province       = form.perm_province.data or None,
                    perm_district       = form.perm_district.data.strip() if form.perm_district.data else None,
                    perm_local_level    = form.perm_local_level.data.strip() if form.perm_local_level.data else None,
                    perm_ward           = form.perm_ward.data.strip() if hasattr(form, 'perm_ward') and form.perm_ward.data else None,
                    perm_tole           = form.perm_tole.data.strip() if hasattr(form, 'perm_tole') and form.perm_tole.data else None,
                    curr_province       = form.curr_province.data,
                    curr_district       = form.curr_district.data.strip(),
                    curr_local_level    = form.curr_local_level.data.strip() if form.curr_local_level.data else None,
                    curr_ward           = form.curr_ward.data.strip() if hasattr(form, 'curr_ward') and form.curr_ward.data else None,
                    curr_tole           = form.curr_tole.data.strip() if hasattr(form, 'curr_tole') and form.curr_tole.data else None,
                    phone1              = form.phone1.data.strip(),
                    phone2              = form.phone2.data.strip() if form.phone2.data and form.phone2.data.strip() else None,
                    blood_group         = form.blood_group.data,
                    last_donation_date  = form.last_donation_date.data,
                    donation_times      = form.donation_times.data or 0,
                    donor_type          = form.donor_type.data,
                    social_link         = form.social_link.data.strip() if form.social_link.data and form.social_link.data.strip() else None,
                    is_active           = True,
                    is_public           = True,
                )
                donor.recalculate_and_save()
                db.session.add(donor)
                db.session.commit()
                
                flash(f'✅ Donor {donor.donor_id} ({donor.full_name}) added successfully!', 'success')
                return redirect(url_for('admin.donors'))
            except Exception as e:
                db.session.rollback()
                flash(f'❌ Failed to add donor: {str(e)}', 'danger')
        else:
            for field_name, errors in form.errors.items():
                label = getattr(form, field_name).label.text if hasattr(form, field_name) and hasattr(getattr(form, field_name), 'label') else field_name
                for error in errors:
                    flash(f'⚠️ {label}: {error}', 'danger')
    
    return render_template('admin/donor_form.html', form=form, action='Add')


@admin_bp.route('/donors/<int:id>/edit', methods=['GET', 'POST'])
@permission_required('manage_donors')
def edit_donor(id):
    donor = Donor.query.get_or_404(id)
    
    if request.method == 'GET':
        form = DonorEditForm(obj=donor)
    else:
        form = DonorEditForm()
    form.record_id.data = donor.id
    
    if request.method == 'POST':
        if form.validate_on_submit():
            try:
                form.populate_obj(donor)
                donor.email = form.email.data.strip() if form.email.data and form.email.data.strip() else None
                donor.phone2 = form.phone2.data.strip() if form.phone2.data and form.phone2.data.strip() else None
                donor.social_link = form.social_link.data.strip() if form.social_link.data and form.social_link.data.strip() else None
                donor.updated_at = datetime.utcnow()
                donor.recalculate_and_save()
                db.session.commit()
                flash(f'✅ Donor {donor.donor_id} ({donor.full_name}) updated successfully!', 'success')
                return redirect(url_for('admin.donors'))
            except Exception as e:
                db.session.rollback()
                flash(f'❌ Failed to update donor: {str(e)}', 'danger')
        else:
            for field_name, errors in form.errors.items():
                label = getattr(form, field_name).label.text if hasattr(form, field_name) and hasattr(getattr(form, field_name), 'label') else field_name
                for error in errors:
                    flash(f'⚠️ {label}: {error}', 'danger')
    
    return render_template('admin/donor_form.html', form=form, donor=donor, action='Edit')


@admin_bp.route('/donors/<int:id>/delete', methods=['POST'])
@permission_required('manage_donors')
def delete_donor(id):
    donor = Donor.query.get_or_404(id)
    db.session.delete(donor)
    db.session.commit()
    flash(f'Donor {donor.donor_id} deleted.', 'warning')
    return redirect(url_for('admin.donors'))


@admin_bp.route('/donors/<int:id>/reset-pin', methods=['POST'])
@permission_required('manage_donors')
def reset_donor_pin(id):
    """Reset a donor's 4-digit login PIN to the default temporary value (1234).
    The donor must change it on next login via their profile.
    """
    donor = Donor.query.get_or_404(id)
    temp_pin = '1234'
    donor.set_pin(temp_pin)
    db.session.commit()

    log_audit_event(
        'RESET_DONOR_PIN', donor.id,
        f'PIN reset for donor {donor.donor_id} ({donor.full_name})',
        actor=current_user.username
    )

    flash(
        f'PIN for donor <strong>{donor.full_name}</strong> ({donor.donor_id}) '
        f'has been reset to <code>1234</code>. '
        f'Please notify the donor to change their PIN on next login.',
        'success'
    )
    return redirect(url_for('admin.donors'))


@admin_bp.route('/donors/<int:id>/toggle-status', methods=['POST'])
@permission_required('manage_donors')
def toggle_donor_status(id):
    donor = Donor.query.get_or_404(id)
    donor.availability_status = 'unavailable' if donor.availability_status == 'available' else 'available'
    db.session.commit()
    return jsonify({'status': donor.availability_status})

@admin_bp.route('/donors/<int:donor_id>/history/<int:history_id>/delete', methods=['POST'])
@permission_required('manage_donors')
def delete_donor_history(donor_id, history_id):
    from app.models import DonorDonationHistory
    donor = Donor.query.get_or_404(donor_id)
    history = DonorDonationHistory.query.filter_by(id=history_id, donor_id=donor.id).first_or_404()
    db.session.delete(history)
    
    # Recalculate donor summary
    donor.donation_times = max(0, (donor.donation_times or 0) - 1)
    donor.total_donations = max(0, (donor.total_donations or 0) - 1)
    
    # Update last donation date
    last_donation = DonorDonationHistory.query.filter_by(donor_id=donor.id).filter(DonorDonationHistory.id != history_id).order_by(DonorDonationHistory.donation_date.desc()).first()
    donor.last_donation_date = last_donation.donation_date if last_donation else None
    
    donor.recalculate_and_save()
    db.session.commit()
    flash('Donation history record deleted.', 'success')
    return redirect(url_for('public.donor_profile', donor_id=donor.donor_id))


# ════════════════════════════════════════════
#   BLOOD BANKS MANAGEMENT
# ════════════════════════════════════════════
@admin_bp.route('/blood-banks')
@permission_required('manage_blood_banks')
def blood_banks():
    page = request.args.get('page', 1, type=int)
    search = request.args.get('q', '')
    
    query = BloodBank.query
    if search:
        query = query.filter(BloodBank.name.ilike(f'%{search}%') | BloodBank.district.ilike(f'%{search}%'))
        
    pagination = query.order_by(BloodBank.created_at.desc()).paginate(page=page, per_page=20, error_out=False)
    return render_template('admin/blood_banks.html', pagination=pagination, search=search)

@admin_bp.route('/blood-banks/create', methods=['GET', 'POST'])
@permission_required('manage_blood_banks')
def create_blood_bank():
    from app.forms import BloodBankForm
    form = BloodBankForm()
    
    if form.validate_on_submit():
        bank = BloodBank(
            name=form.name.data,
            display_name=form.display_name.data,
            hospital_name=form.hospital_name.data,
            branch_type=form.branch_type.data,
            service_type=form.service_type.data,
            province=form.province.data,
            district=form.district.data,
            city=form.city.data,
            contact_number=form.contact_number.data,
            alternate_contact_number=form.alternate_contact_number.data,
            maps_url=form.maps_url.data,
            is_emergency_panel=form.is_emergency_panel.data,
            is_grouped_entry=form.is_grouped_entry.data,
            is_active=form.is_active.data,
            notes=form.notes.data,
            status='active' if form.is_active.data else 'inactive'
        )
        
        db.session.add(bank)
        db.session.commit()
        flash('Blood Bank created successfully!', 'success')
        return redirect(url_for('admin.blood_banks'))
        
    return render_template('admin/blood_bank_form.html', form=form, action='Create')

@admin_bp.route('/blood-banks/<int:id>/edit', methods=['GET', 'POST'])
@permission_required('manage_blood_banks')
def edit_blood_bank(id):
    from app.forms import BloodBankForm
    bank = BloodBank.query.get_or_404(id)
    form = BloodBankForm(obj=bank)
    
    if form.validate_on_submit():
        form.populate_obj(bank)
        bank.status = 'active' if form.is_active.data else 'inactive'
            
        db.session.commit()
        flash('Blood Bank updated successfully!', 'success')
        return redirect(url_for('admin.blood_banks'))
        
    return render_template('admin/blood_bank_form.html', form=form, action='Edit', bank=bank)

@admin_bp.route('/blood-banks/<int:id>/delete', methods=['POST'])
@permission_required('manage_blood_banks')
def delete_blood_bank(id):
    bank = BloodBank.query.get_or_404(id)
    
    # Delete related account (cascades to password_history and login_history)
    if bank.account:
        db.session.delete(bank.account)
    
    # Delete public cache entry
    from app.models import PublicBloodBankCache
    PublicBloodBankCache.query.filter_by(blood_bank_id=id).delete()
    
    db.session.delete(bank)
    db.session.commit()
    flash('Blood Bank and all related data deleted successfully.', 'warning')
    return redirect(url_for('admin.blood_banks'))


# ─── EXPORT: Excel / CSV ───────────────────────────────────────────────────────
@admin_bp.route('/blood-banks/export')
@permission_required('manage_blood_banks')
def export_blood_banks():
    import io, csv
    from flask import make_response
    fmt = request.args.get('format', 'xlsx')  # xlsx or csv

    banks = BloodBank.query.order_by(BloodBank.name).all()

    HEADERS = [
        'name', 'display_name', 'hospital_name', 'branch_type', 'service_type',
        'province', 'district', 'city', 'local_level', 'ward', 'tole',
        'contact_number', 'alternate_contact_number', 'email', 'website',
        'latitude', 'longitude', 'maps_url',
        'is_emergency_panel', 'is_grouped_entry', 'is_active', 'notes',
    ]

    rows = []
    for b in banks:
        rows.append([
            b.name, b.display_name or '', b.hospital_name or '',
            b.branch_type or '', b.service_type or '',
            b.province or '', b.district or '', b.city or '',
            getattr(b, 'local_level', '') or '', getattr(b, 'ward', '') or '', getattr(b, 'tole', '') or '',
            b.contact_number or '', b.alternate_contact_number or '',
            getattr(b, 'email', '') or '', getattr(b, 'website', '') or '',
            str(getattr(b, 'latitude', '') or ''), str(getattr(b, 'longitude', '') or ''),
            b.maps_url or '',
            '1' if b.is_emergency_panel else '0',
            '1' if b.is_grouped_entry else '0',
            '1' if b.is_active else '0',
            b.notes or '',
        ])

    if fmt == 'csv':
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(HEADERS)
        writer.writerows(rows)
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'text/csv; charset=utf-8'
        response.headers['Content-Disposition'] = 'attachment; filename=blood_banks.csv'
        return response

    if fmt == 'pdf':
        try:
            from reportlab.lib.pagesizes import landscape, A4
            from reportlab.lib import colors
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.pdfgen import canvas
        except ImportError:
            flash('reportlab is required for PDF export.', 'danger')
            return redirect(url_for('admin.blood_banks'))

        class NumberedCanvas(canvas.Canvas):
            def __init__(self, *args, **kwargs):
                super().__init__(*args, **kwargs)
                self._saved_page_states = []
            def showPage(self):
                self._saved_page_states.append(dict(self.__dict__))
                self._startPage()
            def save(self):
                num_pages = len(self._saved_page_states)
                for state in self._saved_page_states:
                    self.__dict__.update(state)
                    self.draw_page_number(num_pages)
                    super().showPage()
                super().save()
            def draw_page_number(self, page_count):
                self.saveState()
                self.setFont('Helvetica', 8)
                self.setFillColor(colors.HexColor('#6B7280'))
                self.drawRightString(self._pagesize[0] - 30, 20, f'Page {self._pageNumber} of {page_count}')
                self.drawString(30, 20, 'Raktadata — Nepali Blood Donors Society | Blood Banks Directory')
                self.setStrokeColor(colors.HexColor('#E5E7EB'))
                self.setLineWidth(0.5)
                self.line(30, 32, self._pagesize[0] - 30, 32)
                self.restoreState()

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=landscape(A4), rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=40)
        styles = getSampleStyleSheet()

        title_style = ParagraphStyle('DocTitle', parent=styles['Heading1'], fontName='Helvetica-Bold', fontSize=15, textColor=colors.HexColor('#DC2626'), spaceAfter=4)
        sub_style = ParagraphStyle('DocSub', parent=styles['Normal'], fontName='Helvetica', fontSize=9, textColor=colors.HexColor('#4B5563'), spaceAfter=10)
        cell_style = ParagraphStyle('CellText', parent=styles['Normal'], fontName='Helvetica', fontSize=8, leading=10, textColor=colors.HexColor('#1F2937'))
        header_cell_style = ParagraphStyle('HeaderCellText', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=9, leading=11, textColor=colors.whitesmoke)

        now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        elements = [
            Paragraph('NEPALI BLOOD DONORS SOCIETY — BLOOD BANKS DIRECTORY', title_style),
            Paragraph(f'Exported on: {now_str} | Total Records: {len(banks)}', sub_style),
            Spacer(1, 4)
        ]

        pdf_headers = ['Name', 'Service Type', 'Province', 'District', 'City', 'Phone', 'Emergency', 'Status']
        table_data = [[Paragraph(h, header_cell_style) for h in pdf_headers]]

        for b in banks:
            table_data.append([
                Paragraph(b.display_name or b.name, cell_style),
                Paragraph(b.service_type or 'Blood Bank', cell_style),
                Paragraph(b.province or 'N/A', cell_style),
                Paragraph(b.district or 'N/A', cell_style),
                Paragraph(b.city or 'N/A', cell_style),
                Paragraph(b.contact_number or 'N/A', cell_style),
                Paragraph('Yes' if (b.is_emergency_panel or b.emergency_available) else 'No', cell_style),
                Paragraph('Active' if b.is_active else 'Inactive', cell_style),
            ])

        col_widths = [160, 90, 95, 95, 95, 100, 65, 60]
        t = Table(table_data, colWidths=col_widths, repeatRows=1)
        t_style = [
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#DC2626')),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('BOTTOMPADDING', (0,0), (-1,0), 6),
            ('TOPPADDING', (0,0), (-1,0), 6),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E5E7EB')),
        ]
        for r_i in range(1, len(table_data)):
            bg = colors.HexColor('#F9FAFB') if r_i % 2 == 0 else colors.white
            t_style.append(('BACKGROUND', (0, r_i), (-1, r_i), bg))
            t_style.append(('TOPPADDING', (0, r_i), (-1, r_i), 4))
            t_style.append(('BOTTOMPADDING', (0, r_i), (-1, r_i), 4))

        t.setStyle(TableStyle(t_style))
        elements.append(t)
        doc.build(elements, canvasmaker=NumberedCanvas)

        buf.seek(0)
        response = make_response(buf.read())
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Content-Disposition'] = 'attachment; filename=blood_banks.pdf'
        return response

    # Default: Excel
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        flash('openpyxl is required for Excel export. Please install it.', 'danger')
        return redirect(url_for('admin.blood_banks'))

    wb = Workbook()
    ws = wb.active
    ws.title = 'Blood Banks'

    # Header styling
    header_fill = PatternFill('solid', fgColor='DC2626')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    thin = Side(style='thin', color='D1D5DB')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.append(HEADERS)
    for col_idx, _ in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    ws.row_dimensions[1].height = 30

    # Data rows
    for row_data in rows:
        ws.append(row_data)
        r = ws.max_row
        for col_idx in range(1, len(HEADERS) + 1):
            ws.cell(row=r, column=col_idx).border = border

    # Auto-width
    for col_idx, header in enumerate(HEADERS, 1):
        max_len = max((len(str(row[col_idx-1])) for row in rows), default=0)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(len(header), max_len) + 3, 40)

    # Freeze header row
    ws.freeze_panes = 'A2'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = make_response(buf.read())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = 'attachment; filename=blood_banks.xlsx'
    return response


# ─── DOWNLOAD UPLOAD TEMPLATE ──────────────────────────────────────────────────
@admin_bp.route('/blood-banks/upload-template')
@permission_required('manage_blood_banks')
def blood_bank_upload_template():
    import io
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        flash('openpyxl is required. Please install it.', 'danger')
        return redirect(url_for('admin.blood_banks'))

    HEADERS = [
        'name', 'display_name', 'hospital_name', 'branch_type', 'service_type',
        'province', 'district', 'city', 'local_level', 'ward', 'tole',
        'contact_number', 'alternate_contact_number', 'email', 'website',
        'latitude', 'longitude', 'maps_url',
        'is_emergency_panel', 'is_grouped_entry', 'is_active', 'notes',
    ]
    wb = Workbook()
    ws = wb.active
    ws.title = 'Blood Banks Template'
    ws.append(HEADERS)
    header_fill = PatternFill('solid', fgColor='DC2626')
    header_font = Font(bold=True, color='FFFFFF')
    for col_idx in range(1, len(HEADERS)+1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # One sample row
    ws.append([
        'Example Blood Bank', 'Example BB', 'Teaching Hospital', 'Main', 'Blood Bank',
        'Bagmati Province', 'Kathmandu', 'Kathmandu', 'Kathmandu Metropolitan', '3', 'Baneshwor',
        '01-4780000', '9801234567', 'example@bb.org.np', 'https://example.com',
        '27.7172', '85.3240', 'https://maps.google.com/?q=Kathmandu',
        '1', '0', '1', 'Sample notes here',
    ])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    from flask import make_response
    resp = make_response(buf.read())
    resp.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    resp.headers['Content-Disposition'] = 'attachment; filename=blood_banks_upload_template.xlsx'
    return resp


# ─── BULK UPLOAD ───────────────────────────────────────────────────────────────
@admin_bp.route('/blood-banks/bulk-upload', methods=['POST'])
@permission_required('manage_blood_banks')
def bulk_upload_blood_banks():
    import io, csv as csv_module
    file = request.files.get('bulk_file')
    if not file or not file.filename:
        flash('No file selected.', 'danger')
        return redirect(url_for('admin.blood_banks'))

    filename = file.filename.lower()
    REQUIRED_COLS = {'name', 'province', 'district', 'contact_number', 'service_type', 'maps_url'}

    created = updated = skipped = 0
    errors = []

    try:
        if filename.endswith('.csv'):
            stream = io.StringIO(file.stream.read().decode('utf-8-sig'))
            reader = csv_module.DictReader(stream)
            rows = list(reader)
            headers = set(reader.fieldnames or [])
        elif filename.endswith('.xlsx') or filename.endswith('.xls'):
            try:
                from openpyxl import load_workbook
            except ImportError:
                flash('openpyxl required for Excel upload.', 'danger')
                return redirect(url_for('admin.blood_banks'))
            wb = load_workbook(io.BytesIO(file.stream.read()), read_only=True, data_only=True)
            ws = wb.active
            raw = list(ws.values)
            if not raw:
                flash('Uploaded file is empty.', 'danger')
                return redirect(url_for('admin.blood_banks'))
            header_row = [str(h).strip() if h else '' for h in raw[0]]
            headers = set(header_row)
            rows = [dict(zip(header_row, [str(v).strip() if v is not None else '' for v in row])) for row in raw[1:]]
        else:
            flash('Only .xlsx, .xls, or .csv files are supported.', 'danger')
            return redirect(url_for('admin.blood_banks'))

        # Validate required columns
        missing = REQUIRED_COLS - headers
        if missing:
            flash(f'Missing required columns: {", ".join(sorted(missing))}', 'danger')
            return redirect(url_for('admin.blood_banks'))

        def _bool(val):
            return str(val).strip() in ('1', 'true', 'True', 'yes', 'Yes', 'YES')

        for i, row in enumerate(rows, 2):
            name = str(row.get('name', '')).strip()
            if not name:
                skipped += 1
                continue

            try:
                # Upsert: find by name + district
                district = str(row.get('district', '')).strip()
                bank = BloodBank.query.filter_by(name=name, district=district).first()
                is_new = bank is None
                if is_new:
                    bank = BloodBank()

                bank.name               = name
                bank.display_name       = str(row.get('display_name', '')).strip() or name
                bank.hospital_name      = str(row.get('hospital_name', '')).strip() or None
                bank.branch_type        = str(row.get('branch_type', '')).strip() or None
                bank.service_type       = str(row.get('service_type', '')).strip() or 'Blood Bank'
                bank.province           = str(row.get('province', '')).strip() or None
                bank.district           = district or None
                bank.city               = str(row.get('city', '')).strip() or None
                bank.contact_number     = str(row.get('contact_number', '')).strip() or None
                bank.alternate_contact_number = str(row.get('alternate_contact_number', '')).strip() or None
                bank.maps_url           = str(row.get('maps_url', '')).strip() or None
                bank.notes              = str(row.get('notes', '')).strip() or None
                bank.is_emergency_panel = _bool(row.get('is_emergency_panel', 0))
                bank.is_grouped_entry   = _bool(row.get('is_grouped_entry', 0))
                bank.is_active          = _bool(row.get('is_active', 1))
                bank.status             = 'active' if bank.is_active else 'inactive'

                # Optional extended fields
                for attr in ('local_level', 'ward', 'tole', 'email', 'website'):
                    val = str(row.get(attr, '')).strip()
                    if hasattr(bank, attr):
                        setattr(bank, attr, val or None)
                for attr in ('latitude', 'longitude'):
                    val = str(row.get(attr, '')).strip()
                    if hasattr(bank, attr) and val:
                        try:
                            setattr(bank, attr, float(val))
                        except ValueError:
                            pass

                if is_new:
                    db.session.add(bank)
                    created += 1
                else:
                    updated += 1

            except Exception as e:
                errors.append(f'Row {i} ({name}): {e}')
                skipped += 1

        db.session.commit()
        msg = f'Bulk upload complete — {created} created, {updated} updated, {skipped} skipped.'
        if errors:
            msg += f' Errors: {"; ".join(errors[:5])}'
        flash(msg, 'success' if not errors else 'warning')

    except Exception as e:
        db.session.rollback()
        flash(f'Upload failed: {e}', 'danger')

    return redirect(url_for('admin.blood_banks'))


@admin_bp.route('/blood-banks/<int:id>/generate-account', methods=['POST'])
@permission_required('manage_users')
def generate_blood_bank_account(id):
    bank = BloodBank.query.get_or_404(id)
    
    if hasattr(bank, 'account') and bank.account:
        flash('Account already exists for this blood bank.', 'warning')
        return redirect(url_for('admin.blood_banks'))
        
    try:
        province_code = bank.province[:3].upper() if bank.province else 'UNK'
        district_code = bank.district[:3].upper() if bank.district else 'UNK'
        
        account, raw_password = AuthService.create_blood_bank_account(bank.id, province_code, district_code)
        
        # Provision the tenant database immediately
        from app.services.tenant_service import TenantProvisioningService
        TenantProvisioningService.provision_tenant(bank.id)
        
        log_audit_event('CREATE_BLOOD_BANK_ACCOUNT', bank.id, f'Created account for {bank.name}', actor=current_user.username)
        
        # Enqueue Email Notification
        if bank.email:
            from app.models import Notification, NotificationQueue
            import json
            notif = Notification(
                title='Blood Bank Account Created',
                message=f'Your portal login ID is: {account.login_id}\nYour temporary password is: {raw_password}\nPlease change it upon first login.',
                category='system',
                channel='email'
            )
            db.session.add(notif)
            db.session.flush() # get notif.id
            queue_item = NotificationQueue(
                notification_id=notif.id,
                channel='email',
                payload=json.dumps({'to': bank.email, 'subject': 'Your Nepal Blood Donors Account'})
            )
            db.session.add(queue_item)
            db.session.commit()
            flash(f'Account created and notification queued for {bank.email}', 'success')
        
        # Redirect to a dedicated credentials display page
        return render_template('admin/blood_bank_credentials.html',
                               bank=bank, account=account, raw_password=raw_password)
    except Exception as e:
        db.session.rollback()
        flash(f'Error generating account: {str(e)}', 'danger')
        
    return redirect(url_for('admin.blood_banks'))

@admin_bp.route('/blood-banks/<int:id>/account')
@login_required
@superadmin_required
def view_blood_bank_account(id):
    bank = BloodBank.query.get_or_404(id)
    if not bank.account:
        flash('No account exists for this blood bank.', 'warning')
        return redirect(url_for('admin.blood_banks'))
    return render_template('admin/blood_bank_account_detail.html', bank=bank, account=bank.account)

@admin_bp.route('/blood-banks/<int:id>/account/toggle-lock', methods=['POST'])
@login_required
@superadmin_required
def toggle_blood_bank_account_lock(id):
    bank = BloodBank.query.get_or_404(id)
    if not bank.account:
        flash('No account exists for this blood bank.', 'warning')
        return redirect(url_for('admin.blood_banks'))
    
    action = request.form.get('action', 'lock')
    if action == 'unlock':
        bank.account.is_locked = False
        bank.account.failed_login_attempts = 0
        bank.account.locked_until = None
        flash(f'Account for {bank.resolved_display_name} has been unlocked.', 'success')
        log_audit_event('UNLOCK_BLOOD_BANK_ACCOUNT', bank.id, f'Unlocked account {bank.account.login_id}', actor=current_user.username)
    else:
        bank.account.is_locked = True
        flash(f'Account for {bank.resolved_display_name} has been locked.', 'warning')
        log_audit_event('LOCK_BLOOD_BANK_ACCOUNT', bank.id, f'Locked account {bank.account.login_id}', actor=current_user.username)
    
    db.session.commit()
    return redirect(url_for('admin.view_blood_bank_account', id=bank.id))

@admin_bp.route('/blood-banks/<int:id>/account/reset-password', methods=['POST'])
@login_required
@superadmin_required
def reset_blood_bank_password(id):
    bank = BloodBank.query.get_or_404(id)
    if not bank.account:
        flash('No account exists for this blood bank.', 'warning')
        return redirect(url_for('admin.blood_banks'))
    
    from app.models import BloodBankPasswordHistory
    raw_password = AuthService.generate_secure_password()
    bank.account.set_password(raw_password)
    bank.account.temp_password = raw_password
    bank.account.password_change_required = True
    
    history = BloodBankPasswordHistory(
        # pyrefly: ignore [unexpected-keyword]
        account_id=bank.account.id,
        # pyrefly: ignore [unexpected-keyword]
        password_hash=bank.account.password_hash,
        # pyrefly: ignore [unexpected-keyword]
        created_at=datetime.utcnow()
    )
    db.session.add(history)
    db.session.commit()
    
    log_audit_event('RESET_BLOOD_BANK_PASSWORD', bank.id, f'Password reset for {bank.account.login_id}', actor=current_user.username)
    
    # Enqueue Email Notification
    if bank.email:
        from app.models import Notification, NotificationQueue
        import json
        notif = Notification(
            title='Blood Bank Password Reset',
            message=f'Your portal password has been reset.\nYour new temporary password is: {raw_password}\nPlease change it upon login.',
            category='system',
            channel='email'
        )
        db.session.add(notif)
        db.session.flush() # get notif.id
        queue_item = NotificationQueue(
            notification_id=notif.id,
            channel='email',
            payload=json.dumps({'to': bank.email, 'subject': 'Nepal Blood Donors - Password Reset'})
        )
        db.session.add(queue_item)
        flash(f'Password reset and notification queued for {bank.email}', 'success')
        
    db.session.commit()
    
    # Show the new credentials page
    return render_template('admin/blood_bank_credentials.html',
                           bank=bank, account=bank.account, raw_password=raw_password)


# ════════════════════════════════════════════
#   BLOOD REQUEST MANAGEMENT
# ════════════════════════════════════════════
@admin_bp.route('/requests')
@permission_required('manage_requests')
def blood_requests():
    page        = request.args.get('page', 1, type=int)
    status      = request.args.get('status', '')
    blood_group = request.args.get('bg', '')
    
    query = BloodRequest.query
    
    if status:
        query = query.filter_by(status=status)
    if blood_group:
        query = query.filter_by(blood_group=blood_group)
    
    pagination = paginate_query(
        query.order_by(
            BloodRequest.is_emergency.desc(),
            desc(BloodRequest.created_at)
        ), page, 20
    )
    
    counts = {
        'all': BloodRequest.query.count(),
        'active': BloodRequest.query.filter_by(status='active').count(),
        'fulfilled': BloodRequest.query.filter_by(status='fulfilled').count(),
        'closed': BloodRequest.query.filter_by(status='closed').count(),
    }
    
    return render_template('admin/requests.html',
        pagination=pagination,
        counts=counts,
        selected_status=status,
        selected_bg=blood_group,
        blood_groups=['A+', 'A-', 'B+', 'B-', 'AB+', 'AB-', 'O+', 'O-'],
    )


@admin_bp.route('/requests/<int:id>/status/<string:new_status>', methods=['POST'])
@permission_required('manage_requests')
def update_request_status(id, new_status):
    req = BloodRequest.query.get_or_404(id)
    if new_status in ('active', 'fulfilled', 'closed'):
        req.status = new_status
        db.session.commit()
        flash(f'Request {req.request_id} marked as {new_status}.', 'success')
    return redirect(url_for('admin.blood_requests'))


@admin_bp.route('/requests/<int:id>/verify-paper/<string:action>', methods=['POST'])
@permission_required('manage_users')
def verify_hospital_paper(id, action):
    req = BloodRequest.query.get_or_404(id)
    if action == 'verify':
        req.hospital_paper_verified = True
        flash(f'Hospital paper for request {req.request_id} approved.', 'success')
    elif action == 'reject':
        req.hospital_paper_verified = False
        flash(f'Hospital paper for request {req.request_id} rejected.', 'warning')
    db.session.commit()
    return redirect(url_for('admin.blood_requests'))


@admin_bp.route('/requests/<int:id>/delete', methods=['POST'])
@permission_required('manage_requests')
def delete_request(id):
    req = BloodRequest.query.get_or_404(id)
    db.session.delete(req)
    db.session.commit()
    flash(f'Request {req.request_id} deleted.', 'warning')
    return redirect(url_for('admin.blood_requests'))


# ------------------------------------------------------------------------------
# NOTIFICATION DELIVERY LOGS
# ------------------------------------------------------------------------------
@admin_bp.route('/delivery_logs')
@login_required
@permission_required('manage_users')
def delivery_logs():
    page = request.args.get('page', 1, type=int)
    from app.models import NotificationDeliveryLog
    logs = NotificationDeliveryLog.query.order_by(NotificationDeliveryLog.created_at.desc()).paginate(page=page, per_page=50)
    return render_template('admin/delivery_logs.html', logs=logs)


# ════════════════════════════════════════════
#   NEWS MANAGEMENT
# ════════════════════════════════════════════
@admin_bp.route('/news')
@permission_required('manage_news')
def news():
    page = request.args.get('page', 1, type=int)
    pagination = paginate_query(
        News.query.order_by(desc(News.created_at)), page, 15
    )
    return render_template('admin/news.html', pagination=pagination)


@admin_bp.route('/news/add', methods=['GET', 'POST'])
@permission_required('manage_news')
def add_news():
    form = NewsForm()
    
    if form.validate_on_submit():
        image_file = None
        if form.featured_image.data and form.featured_image.data.filename:
            image_file = save_image(form.featured_image.data, 'news')
        
        post = News(
            title       = form.title.data.strip(),
            short_desc  = form.short_desc.data.strip(),
            content     = sanitize_html(form.content.data),
            category    = form.category.data,
            author      = form.author.data.strip(),
            tags        = form.tags.data.strip() if form.tags.data else None,
            featured_image = image_file,
            is_published = form.is_published.data,
        )
        db.session.add(post)
        db.session.commit()
        
        flash('✅ News post created!', 'success')
        return redirect(url_for('admin.news'))
    
    return render_template('admin/news_form.html', form=form, action='Add')


@admin_bp.route('/news/<int:id>/edit', methods=['GET', 'POST'])
@permission_required('manage_news')
def edit_news(id):
    post = News.query.get_or_404(id)
    form = NewsForm(obj=post)
    
    if form.validate_on_submit():
        if form.featured_image.data and form.featured_image.data.filename:
            delete_file(post.featured_image, 'news')
            post.featured_image = save_image(form.featured_image.data, 'news')
        
        post.title      = form.title.data.strip()
        post.short_desc = form.short_desc.data.strip()
        post.content    = sanitize_html(form.content.data)
        post.category   = form.category.data
        post.author     = form.author.data.strip()
        post.tags       = form.tags.data.strip() if form.tags.data else None
        post.is_published = form.is_published.data
        post.updated_at = datetime.utcnow()
        
        db.session.commit()
        flash('✅ News post updated!', 'success')
        return redirect(url_for('admin.news'))
    
    return render_template('admin/news_form.html', form=form, post=post, action='Edit')


@admin_bp.route('/news/<int:id>/delete', methods=['POST'])
@permission_required('manage_news')
def delete_news(id):
    post = News.query.get_or_404(id)
    delete_file(post.featured_image, 'news')
    db.session.delete(post)
    db.session.commit()
    flash('News post deleted.', 'warning')
    return redirect(url_for('admin.news'))


# ════════════════════════════════════════════
#   NOTICE MANAGEMENT
# ════════════════════════════════════════════
@admin_bp.route('/notices')
@permission_required('manage_notices')
def notices():
    page = request.args.get('page', 1, type=int)
    pagination = paginate_query(
        Notice.query.order_by(Notice.priority.desc(), desc(Notice.published_date)), page, 15
    )
    return render_template('admin/notices.html', pagination=pagination)


@admin_bp.route('/notices/add', methods=['GET', 'POST'])
@permission_required('manage_notices')
def add_notice():
    form = NoticeForm()
    
    if form.validate_on_submit():
        file_name, file_ext = None, None
        if form.attachment.data and form.attachment.data.filename:
            file_name, file_ext = save_file(form.attachment.data, 'notices')
        
        notice = Notice(
            # pyrefly: ignore [unexpected-keyword]
            title           = form.title.data.strip(),
            # pyrefly: ignore [unexpected-keyword]
            content         = form.content.data.strip(),
            # pyrefly: ignore [unexpected-keyword]
            expiry_date     = datetime.combine(form.expiry_date.data, datetime.min.time()) if form.expiry_date.data else None,
            # pyrefly: ignore [unexpected-keyword]
            priority        = int(form.priority.data),
            # pyrefly: ignore [unexpected-keyword]
            attachment      = file_name,
            # pyrefly: ignore [unexpected-keyword]
            attachment_type = file_ext,
            # pyrefly: ignore [unexpected-keyword]
            is_active       = form.is_active.data,
        )
        db.session.add(notice)
        db.session.commit()
        
        flash('✅ Notice published!', 'success')
        return redirect(url_for('admin.notices'))
    
    return render_template('admin/notice_form.html', form=form, action='Add')


@admin_bp.route('/notices/<int:id>/edit', methods=['GET', 'POST'])
@permission_required('manage_notices')
def edit_notice(id):
    notice = Notice.query.get_or_404(id)
    if request.method == 'GET':
        form = NoticeForm(obj=notice)
        if notice.expiry_date and isinstance(notice.expiry_date, datetime):
            form.expiry_date.data = notice.expiry_date.date()
    else:
        form = NoticeForm()
    
    if form.validate_on_submit():
        if form.attachment.data and form.attachment.data.filename:
            if notice.attachment:
                delete_file(notice.attachment, 'notices')
            file_name, file_ext = save_file(form.attachment.data, 'notices')
            notice.attachment = file_name
            notice.attachment_type = file_ext
            
        notice.title = form.title.data.strip()
        notice.content = form.content.data.strip()
        if form.expiry_date.data:
            ed = form.expiry_date.data
            notice.expiry_date = datetime.combine(ed, datetime.min.time()) if not isinstance(ed, datetime) else ed
        else:
            notice.expiry_date = None
        notice.priority = int(form.priority.data) if form.priority.data is not None else 0
        notice.is_active = form.is_active.data
        notice.updated_at = datetime.utcnow()
        
        db.session.commit()
        flash('✅ Notice updated successfully!', 'success')
        return redirect(url_for('admin.notices'))
        
    return render_template('admin/notice_form.html', form=form, notice=notice, action='Edit')


@admin_bp.route('/notices/<int:id>/delete', methods=['POST'])
@permission_required('manage_notices')
def delete_notice(id):
    notice = Notice.query.get_or_404(id)
    delete_file(notice.attachment, 'notices')
    db.session.delete(notice)
    db.session.commit()
    flash('Notice deleted.', 'warning')
    return redirect(url_for('admin.notices'))


@admin_bp.route('/notices/<int:id>/toggle', methods=['POST'])
@permission_required('manage_notices')
def toggle_notice(id):
    notice = Notice.query.get_or_404(id)
    notice.is_active = not notice.is_active
    db.session.commit()
    return jsonify({'is_active': notice.is_active})


# ════════════════════════════════════════════
#   ADVERTISEMENT MANAGEMENT
# ════════════════════════════════════════════
@admin_bp.route('/advertisements')
@permission_required('manage_ads')
def advertisements():
    page = request.args.get('page', 1, type=int)
    pagination = paginate_query(
        Advertisement.query.order_by(desc(Advertisement.created_at)), page, 15
    )
    
    # Monthly click report (Database dialect aware: PostgreSQL vs SQLite)
    try:
        bind = db.session.get_bind()
        dialect_name = bind.dialect.name if bind else 'sqlite'
        if dialect_name == 'postgresql':
            month_col = func.to_char(Advertisement.created_at, 'YYYY-MM').label('month')
            monthly_clicks = db.session.query(
                month_col,
                func.sum(Advertisement.clicks).label('total_clicks'),
                func.sum(Advertisement.impressions).label('total_impressions'),
            ).group_by(month_col).order_by(desc(month_col)).limit(6).all()
        else:
            monthly_clicks = db.session.query(
                func.strftime('%Y-%m', Advertisement.created_at).label('month'),
                func.sum(Advertisement.clicks).label('total_clicks'),
                func.sum(Advertisement.impressions).label('total_impressions'),
            ).group_by('month').order_by(desc('month')).limit(6).all()
    except Exception:
        monthly_clicks = []
    
    return render_template('admin/advertisements.html',
        pagination=pagination,
        monthly_clicks=monthly_clicks,
    )


@admin_bp.route('/advertisements/add', methods=['GET', 'POST'])
@permission_required('manage_users')
def add_advertisement():
    form = AdvertisementForm()
    
    if form.validate_on_submit():
        if not form.image.data or not form.image.data.filename:
            flash('Banner image is required.', 'danger')
            return render_template('admin/ad_form.html', form=form, action='Add')
        
        image_file = save_image(form.image.data, 'ads', max_width=800, max_height=600)
        
        ad = Advertisement(
            # pyrefly: ignore [unexpected-keyword]
            title       = form.title.data.strip(),
            # pyrefly: ignore [unexpected-keyword]
            description = form.description.data.strip() if form.description.data else None,
            # pyrefly: ignore [unexpected-keyword]
            image       = image_file,
            # pyrefly: ignore [unexpected-keyword]
            redirect_url= form.redirect_url.data.strip() if form.redirect_url.data else None,
            # pyrefly: ignore [unexpected-keyword]
            ad_type     = form.ad_type.data,
            # pyrefly: ignore [unexpected-keyword]
            start_date  = datetime.combine(form.start_date.data, datetime.min.time()),
            # pyrefly: ignore [unexpected-keyword]
            end_date    = datetime.combine(form.end_date.data, datetime.max.time()),
            # pyrefly: ignore [unexpected-keyword]
            is_active   = form.is_active.data,
        )
        db.session.add(ad)
        db.session.commit()
        
        flash('✅ Advertisement created!', 'success')
        return redirect(url_for('admin.advertisements'))
    
    return render_template('admin/ad_form.html', form=form, action='Add')


@admin_bp.route('/advertisements/<int:id>/toggle', methods=['POST'])
@permission_required('manage_users')
def toggle_ad(id):
    ad = Advertisement.query.get_or_404(id)
    ad.is_active = not ad.is_active
    db.session.commit()
    flash(f"Advertisement {'activated' if ad.is_active else 'deactivated'}.", 'success')
    return redirect(url_for('admin.advertisements'))


@admin_bp.route('/advertisements/<int:id>/delete', methods=['POST'])
@permission_required('manage_users')
def delete_advertisement(id):
    ad = Advertisement.query.get_or_404(id)
    delete_file(ad.image, 'ads')
    db.session.delete(ad)
    db.session.commit()
    flash('Advertisement deleted.', 'warning')
    return redirect(url_for('admin.advertisements'))


# ════════════════════════════════════════════
#   CONTACTS
# ════════════════════════════════════════════
@admin_bp.route('/contacts')
@permission_required('moderate_content')
def contacts():
    page = request.args.get('page', 1, type=int)
    pagination = paginate_query(
        Contact.query.order_by(Contact.is_read.asc(), desc(Contact.created_at)), page, 20
    )
    return render_template('admin/contacts.html', pagination=pagination)


@admin_bp.route('/contacts/<int:id>/read', methods=['POST'])
@permission_required('moderate_content')
def mark_contact_read(id):
    msg = Contact.query.get_or_404(id)
    msg.is_read = True
    db.session.commit()
    
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return jsonify({'ok': True})
        
    flash('Message marked as read.', 'success')
    return redirect(url_for('admin.contacts'))


# ════════════════════════════════════════════
#   SUCCESS STORIES MANAGEMENT (Admin Panel)
# ════════════════════════════════════════════
@admin_bp.route('/success-stories')
@permission_required('manage_users')
def success_stories():
    """एडमिन ड्यासबोर्ड भित्र सबै सफलताका कथाहरू सूचीकृत गर्ने मुख्य व्यवस्थापन राउट"""
    page = request.args.get('page', 1, type=int)
    pagination = paginate_query(
        SuccessStory.query.order_by(desc(SuccessStory.created_at)), page, 15
    )
    return render_template('admin/success_stories.html', pagination=pagination)


@admin_bp.route('/success-stories/<int:id>/status/<string:new_status>', methods=['POST'])
@permission_required('manage_success_stories')
def update_story_status(id, new_status):
    """सफलताका कथाहरूको स्थिति (Pending, Approved, Rejected) परिमार्जन गर्ने"""
    story = SuccessStory.query.get_or_404(id)
    if new_status in ['pending', 'approved', 'rejected', 'hidden']:
        story.status = new_status
        db.session.commit()
        flash(f'Story status updated to {new_status}.', 'success')
    else:
        flash('Invalid status.', 'danger')
    return redirect(url_for('admin.success_stories'))

@admin_bp.route('/success-stories/<int:id>/delete', methods=['POST'])
@permission_required('manage_success_stories')
def delete_success_story(id):
    """एडमिन प्यानल र सर्भर स्टोरेज दुवैबाट कथा सुरक्षित रूपमा डिलिट गर्ने राउट"""
    story = SuccessStory.query.get_or_404(id)
    
    # यदि कथासँग कुनै अपलोड गरिएको तस्बिर छ भने त्यसलाई पनि सर्भरबाट सधैँका लागि सफा गर्ने
    if story.image_file:
        delete_file(story.image_file, 'stories')
        
    db.session.delete(story)
    db.session.commit()
    
    flash(f'⚠️ Success story by "{story.author_name}" has been permanently deleted.', 'warning')
    return redirect(url_for('admin.success_stories'))




# ════════════════════════════════════════════
#   STAFF MANAGEMENT
# ════════════════════════════════════════════
@admin_bp.route('/staff')
@permission_required('manage_staff')
def staff():
    page = request.args.get('page', 1, type=int)
    pagination = paginate_query(
        StaffMember.query.order_by(desc(StaffMember.created_at)), page, 15
    )
    return render_template('admin/staff.html', pagination=pagination)


@admin_bp.route('/staff/add', methods=['GET', 'POST'])
@permission_required('manage_staff')
def add_staff():
    form = StaffMemberForm()
    if form.validate_on_submit():
        photo_file = None
        if form.profile_photo.data and form.profile_photo.data.filename:
            photo_file = save_image(form.profile_photo.data, 'staff')
        
        member = StaffMember(
            full_name=form.full_name.data.strip(),
            designation=form.designation.data.strip(),
            email=form.email.data.strip() if form.email.data else None,
            contact_number=form.contact_number.data.strip() if form.contact_number.data else None,
            profile_photo=photo_file,
            province=form.province.data or None,
            district=form.district.data or None,
            local_level=form.local_level.data or None,
            ward_number=form.ward_number.data or None,
            tole=form.tole.data or None,
            is_active=form.is_active.data
        )
        db.session.add(member)
        db.session.commit()
        flash('✅ Staff member added successfully!', 'success')
        return redirect(url_for('admin.staff'))
    
    return render_template('admin/staff_form.html', form=form, action='Add')


@admin_bp.route('/staff/<int:id>/edit', methods=['GET', 'POST'])
@permission_required('manage_staff')
def edit_staff(id):
    member = StaffMember.query.get_or_404(id)
    form = StaffMemberForm(obj=member)
    
    if form.validate_on_submit():
        if form.profile_photo.data and form.profile_photo.data.filename:
            if member.profile_photo:
                delete_file(member.profile_photo, 'staff')
            member.profile_photo = save_image(form.profile_photo.data, 'staff')
            
        member.full_name = form.full_name.data.strip()
        member.designation = form.designation.data.strip()
        member.email = form.email.data.strip() if form.email.data else None
        member.contact_number = form.contact_number.data.strip() if form.contact_number.data else None
        member.province = form.province.data or None
        member.district = form.district.data or None
        member.local_level = form.local_level.data or None
        member.ward_number = form.ward_number.data or None
        member.tole = form.tole.data or None
        member.is_active = form.is_active.data
        
        db.session.commit()
        flash('✅ Staff member updated successfully!', 'success')
        return redirect(url_for('admin.staff'))
        
    return render_template('admin/staff_form.html', form=form, member=member, action='Edit')


@admin_bp.route('/staff/<int:id>/delete', methods=['POST'])
@permission_required('manage_staff')
def delete_staff(id):
    member = StaffMember.query.get_or_404(id)
    if member.profile_photo:
        delete_file(member.profile_photo, 'staff')
    db.session.delete(member)
    db.session.commit()
    flash('Staff member deleted.', 'warning')
    return redirect(url_for('admin.staff'))


# ════════════════════════════════════════════
#   PARTNER MANAGEMENT
# ════════════════════════════════════════════
@admin_bp.route('/partners')
@permission_required('manage_partners')
def partners():
    page = request.args.get('page', 1, type=int)
    pagination = paginate_query(
        Partner.query.order_by(desc(Partner.created_at)), page, 15
    )
    return render_template('admin/partners.html', pagination=pagination)


@admin_bp.route('/partners/add', methods=['GET', 'POST'])
@permission_required('manage_partners')
def add_partner():
    form = PartnerForm()
    if form.validate_on_submit():
        logo_file = None
        if form.logo_file.data and form.logo_file.data.filename:
            logo_file = save_image(form.logo_file.data, 'partners')
            
        partner = Partner(
            partner_name=form.partner_name.data.strip(),
            description=form.description.data.strip() if form.description.data else None,
            website_url=form.website_url.data.strip() if form.website_url.data else None,
            email=form.email.data.strip() if form.email.data else None,
            contact_number=form.contact_number.data.strip() if form.contact_number.data else None,
            address=form.address.data.strip() if form.address.data else None,
            logo_file=logo_file,
            is_active=form.is_active.data
        )
        db.session.add(partner)
        db.session.commit()
        flash('✅ Partner added successfully!', 'success')
        return redirect(url_for('admin.partners'))
        
    return render_template('admin/partner_form.html', form=form, action='Add')


@admin_bp.route('/partners/<int:id>/edit', methods=['GET', 'POST'])
@permission_required('manage_partners')
def edit_partner(id):
    partner = Partner.query.get_or_404(id)
    form = PartnerForm(obj=partner)
    
    if form.validate_on_submit():
        if form.logo_file.data and form.logo_file.data.filename:
            if partner.logo_file:
                delete_file(partner.logo_file, 'partners')
            partner.logo_file = save_image(form.logo_file.data, 'partners')
            
        partner.partner_name = form.partner_name.data.strip()
        partner.description = form.description.data.strip() if form.description.data else None
        partner.website_url = form.website_url.data.strip() if form.website_url.data else None
        partner.email = form.email.data.strip() if form.email.data else None
        partner.contact_number = form.contact_number.data.strip() if form.contact_number.data else None
        partner.address = form.address.data.strip() if form.address.data else None
        partner.is_active = form.is_active.data
        
        db.session.commit()
        flash('✅ Partner updated successfully!', 'success')
        return redirect(url_for('admin.partners'))
        
    return render_template('admin/partner_form.html', form=form, partner=partner, action='Edit')


@admin_bp.route('/partners/<int:id>/delete', methods=['POST'])
@permission_required('manage_partners')
def delete_partner(id):
    partner = Partner.query.get_or_404(id)
    if partner.logo_file:
        delete_file(partner.logo_file, 'partners')
    db.session.delete(partner)
    db.session.commit()
    flash('Partner deleted.', 'warning')
    return redirect(url_for('admin.partners'))


# ════════════════════════════════════════════
#   DATA QUALITY & ML OPS ENGINE
# ════════════════════════════════════════════
@admin_bp.route('/data-quality')
@permission_required('manage_users')
def data_quality():
    # Calculate Data Quality Metrics
    total_donors = Donor.query.count()
    if total_donors == 0:
        return render_template('admin/data_quality.html', total_donors=0)
        
    # Profile completeness check
    missing_email = Donor.query.filter((Donor.email == None) | (Donor.email == '')).count()
    missing_phone2 = Donor.query.filter((Donor.phone2 == None) | (Donor.phone2 == '')).count()
    missing_last_donation = Donor.query.filter(Donor.last_donation_date == None).count()
    missing_social = Donor.query.filter((Donor.social_link == None) | (Donor.social_link == '')).count()
    missing_perm_address = Donor.query.filter((Donor.perm_district == None) | (Donor.perm_district == '')).count()
    
    completeness_score = int(100 - ((missing_email + missing_last_donation/2 + missing_perm_address) / (total_donors * 3) * 100))
    completeness_score = max(0, min(100, completeness_score))
    
    # Blood Group Supply vs Demand Imbalance
    # Ratios of Available Donors / Active Requests
    groups = ['A+', 'A-', 'B+', 'B-', 'AB+', 'AB-', 'O+', 'O-']
    group_imbalances = []
    
    for g in groups:
        supply = Donor.query.filter_by(blood_group=g, availability_status='available').count()
        demand = BloodRequest.query.filter_by(blood_group=g, status='active').count()
        
        # Determine risk level accurately
        if demand > 0 and supply == 0:
            status = 'Critical Shortage'
            badge = 'danger'
        elif demand > 0 and supply < demand:
            status = 'Critical Deficit'
            badge = 'danger'
        elif demand > 0 and (supply / demand) < 2:
            status = 'High Deficit'
            badge = 'warning'
        elif supply == 0:
            status = 'No Active Donors'
            badge = 'secondary'
        elif supply <= 2:
            status = 'Low Reserve'
            badge = 'warning'
        elif demand > 0 and (supply / demand) >= 2:
            status = 'Optimal'
            badge = 'success'
        else:
            status = 'Healthy Supply'
            badge = 'success'
            
        group_imbalances.append({
            'group': g,
            'supply': supply,
            'demand': demand,
            'status': status,
            'badge': badge
        })
        
    # Regional Imbalance (Top active districts)
    districts = db.session.query(Donor.curr_district, func.count(Donor.id)).group_by(Donor.curr_district).order_by(desc(func.count(Donor.id))).limit(5).all()
    
    # Age group distribution
    age_18_30 = Donor.query.filter(Donor.age >= 18, Donor.age <= 30).count()
    age_31_45 = Donor.query.filter(Donor.age >= 31, Donor.age <= 45).count()
    age_46_65 = Donor.query.filter(Donor.age >= 46, Donor.age <= 65).count()
    
    age_dist = {
        'young': int((age_18_30 / total_donors) * 100) if total_donors > 0 else 0,
        'adult': int((age_31_45 / total_donors) * 100) if total_donors > 0 else 0,
        'senior': int((age_46_65 / total_donors) * 100) if total_donors > 0 else 0
    }
    
    # Duplicate detection (Verified matching contact or high name similarity without family-name false positives)
    all_donors = Donor.query.all()
    potential_duplicates = []
    from difflib import SequenceMatcher
    
    for i in range(len(all_donors)):
        for j in range(i + 1, len(all_donors)):
            d1 = all_donors[i]
            d2 = all_donors[j]
            if d1.id == d2.id:
                continue

            # Check matching secondary phone
            if (d1.phone2 and d2.phone1 and d1.phone2 == d2.phone1) or (d2.phone2 and d1.phone1 and d2.phone2 == d1.phone1):
                potential_duplicates.append({
                    'donor1': d1,
                    'donor2': d2,
                    'similarity': 95,
                    'reason': 'Shared Secondary Phone'
                })
                continue

            # Check matching email
            if d1.email and d2.email and d1.email.strip().lower() == d2.email.strip().lower():
                potential_duplicates.append({
                    'donor1': d1,
                    'donor2': d2,
                    'similarity': 99,
                    'reason': 'Matching Email'
                })
                continue
                
            name1 = d1.full_name.strip().lower()
            name2 = d2.full_name.strip().lower()
            
            # Exact full name match
            if name1 == name2:
                sim = 100 if (d1.blood_group == d2.blood_group and d1.curr_district == d2.curr_district) else 90
                potential_duplicates.append({
                    'donor1': d1,
                    'donor2': d2,
                    'similarity': sim,
                    'reason': 'Identical Full Name'
                })
                continue
                
            # Tokenized comparison to prevent family-member false positives (e.g. Prajwal Timsina vs Ujwal Timsina)
            tokens1 = name1.split()
            tokens2 = name2.split()
            
            if len(tokens1) >= 2 and len(tokens2) >= 2:
                first1, last1 = tokens1[0], tokens1[-1]
                first2, last2 = tokens2[0], tokens2[-1]
                
                first_sim = SequenceMatcher(None, first1, first2).ratio()
                last_sim = SequenceMatcher(None, last1, last2).ratio()
                
                # Both first name and last name must have high similarity (e.g. typos, spelling variations)
                if first_sim >= 0.82 and last_sim >= 0.82:
                    overall_sim = SequenceMatcher(None, name1, name2).ratio()
                    if overall_sim >= 0.88:
                        potential_duplicates.append({
                            'donor1': d1,
                            'donor2': d2,
                            'similarity': int(overall_sim * 100),
                            'reason': 'High Name Similarity'
                        })
                    
    return render_template('admin/data_quality.html',
        total_donors=total_donors,
        missing_email=missing_email,
        missing_phone2=missing_phone2,
        missing_last_donation=missing_last_donation,
        missing_social=missing_social,
        missing_perm_address=missing_perm_address,
        completeness_score=completeness_score,
        group_imbalances=group_imbalances,
        districts=districts,
        age_dist=age_dist,
        potential_duplicates=potential_duplicates
    )
@admin_bp.route('/notifications-dashboard')
@login_required
def notifications_dashboard():
    from app.models import NotificationQueue, NotificationDeliveryLog
    from sqlalchemy import func
    
    # Aggregates for Chart.js
    channel_stats = db.session.query(
        NotificationDeliveryLog.channel,
        func.count(NotificationDeliveryLog.id).label('total'),
        func.sum(db.case((NotificationDeliveryLog.status == 'sent', 1), else_=0)).label('success'),
        func.sum(db.case((NotificationDeliveryLog.status == 'failed', 1), else_=0)).label('failed')
    ).group_by(NotificationDeliveryLog.channel).all()
    
    queue_stats = db.session.query(
        NotificationQueue.status, func.count(NotificationQueue.id)
    ).group_by(NotificationQueue.status).all()
    
    recent_logs = NotificationDeliveryLog.query.order_by(NotificationDeliveryLog.created_at.desc()).limit(50).all()
    
    return render_template(
        'admin/notifications_dashboard.html',
        channel_stats=channel_stats,
        queue_stats=queue_stats,
        recent_logs=recent_logs
    )

# ════════════════════════════════════════════
#   ADMIN USERS MANAGEMENT (SUPERADMIN ONLY)
# ════════════════════════════════════════════
@admin_bp.route('/users')
@superadmin_required
def users():
    page = request.args.get('page', 1, type=int)
    pagination = User.query.order_by(User.created_at.desc()).paginate(page=page, per_page=20)
    return render_template('admin/users.html', pagination=pagination)

@admin_bp.route('/users/add', methods=['GET', 'POST'])
@superadmin_required
def add_user():

    form = AdminUserForm()
    
    if form.validate_on_submit():
        user = User(
            username=form.username.data,
            email=form.email.data,
            full_name=form.full_name.data,
            role=form.role.data,
            is_active=form.is_active.data
        )
        password = form.password.data if form.password.data else 'admin123'
        user.set_password(password)
        db.session.add(user)
        db.session.commit()
        flash('Admin user created successfully.', 'success')
        return redirect(url_for('admin.users'))
        
    return render_template('admin/user_form.html', form=form, action='Add')

@admin_bp.route('/users/<int:id>/edit', methods=['GET', 'POST'])
@superadmin_required
def edit_user(id):

    user = User.query.get_or_404(id)
    if user.id == current_user.id:
        flash("You cannot edit your own role here. Use profile settings.", "warning")
        return redirect(url_for('admin.users'))
        
    form = AdminUserForm(obj=user)
    
    if form.validate_on_submit():
        form.populate_obj(user)
        if form.password.data:
            user.set_password(form.password.data)
        db.session.commit()
        flash('Admin user updated successfully!', 'success')
        return redirect(url_for('admin.users'))
        
    return render_template('admin/user_form.html', form=form, action='Edit', user=user)

@admin_bp.route('/users/<int:id>/delete', methods=['POST'])
@superadmin_required
def delete_user(id):
    user = User.query.get_or_404(id)
    if user.id == current_user.id:
        flash("You cannot delete yourself.", "danger")
        return redirect(url_for('admin.users'))
        
    db.session.delete(user)
    db.session.commit()
    flash('Admin user deleted successfully.', 'warning')
    return redirect(url_for('admin.users'))


@admin_bp.route('/users/<int:id>/reset-password', methods=['POST'])
@superadmin_required
def reset_admin_password(id):
    """Admin-side password reset for system admin/moderator accounts.
    Generates a secure temporary password, forces change on next login.
    """
    user = User.query.get_or_404(id)
    if user.id == current_user.id:
        flash("Use your profile settings to change your own password.", "warning")
        return redirect(url_for('admin.users'))

    new_password = AuthService.generate_secure_password()
    user.set_password(new_password)
    db.session.commit()

    log_audit_event(
        'RESET_ADMIN_PASSWORD', user.id,
        f'Admin password reset for user: {user.username}',
        actor=current_user.username
    )

    flash(
        f'Password for <strong>{user.username}</strong> has been reset. '
        f'New temporary password: <code>{new_password}</code> — '
        f'Please share it securely and ask them to change it on next login.',
        'success'
    )
    return redirect(url_for('admin.users'))


# ════════════════════════════════════════════
#   VOLUNTEER APPROVALS MANAGEMENT
# ════════════════════════════════════════════
@admin_bp.route('/volunteers')
@permission_required('manage_volunteer_approvals')
def volunteers():
    page = request.args.get('page', 1, type=int)
    search = request.args.get('search', '').strip()
    
    query = Volunteer.query
    if search:
        query = query.filter(
            or_(
                Volunteer.full_name.like(f'%{search}%'),
                Volunteer.email.like(f'%{search}%'),
                Volunteer.phone1.like(f'%{search}%')
            )
        )
    
    pagination = paginate_query(
        query.order_by(desc(Volunteer.created_at)), page, 15
    )
    return render_template('admin/volunteers.html', pagination=pagination, search=search)

@admin_bp.route('/volunteers/<int:id>/approve', methods=['POST'])
@permission_required('manage_volunteer_approvals')
def approve_volunteer(id):
    volunteer = Volunteer.query.get_or_404(id)
    volunteer.is_approved = True
    db.session.commit()
    flash(f'Volunteer {volunteer.full_name} has been approved!', 'success')
    return redirect(url_for('admin.volunteers'))

@admin_bp.route('/volunteers/<int:id>/toggle-active', methods=['POST'])
@permission_required('manage_volunteer_approvals')
def toggle_volunteer_active(id):
    volunteer = Volunteer.query.get_or_404(id)
    volunteer.is_active = not volunteer.is_active
    db.session.commit()
    status = 'activated' if volunteer.is_active else 'deactivated'
    flash(f'Volunteer {volunteer.full_name} has been {status}!', 'info')
    return redirect(url_for('admin.volunteers'))

@admin_bp.route('/volunteers/<int:id>/delete', methods=['POST'])
@permission_required('manage_volunteer_approvals')
def delete_volunteer(id):
    volunteer = Volunteer.query.get_or_404(id)
    db.session.delete(volunteer)
    db.session.commit()
    flash(f'Volunteer {volunteer.full_name} deleted successfully.', 'warning')
    return redirect(url_for('admin.volunteers'))


@admin_bp.context_processor
def inject_admin_globals():
    from app.rbac import has_permission as check_permission
    if current_user.is_authenticated:
        from app.models import Donor, BloodRequest, Contact, Volunteer
        try:
            donor_count = Donor.query.count()
            active_req_count = BloodRequest.query.filter_by(status='active').count()
            unread_count = Contact.query.filter_by(is_read=False).count()
            pending_volunteers = Volunteer.query.filter_by(is_approved=False).count()
            return dict(
                donor_count=donor_count,
                active_req_count=active_req_count,
                unread_count=unread_count,
                pending_volunteers_count=pending_volunteers,
                has_permission=lambda perm: check_permission(current_user, perm)
            )
        except Exception:
            pass
    return dict(
        has_permission=lambda perm: False
    )



